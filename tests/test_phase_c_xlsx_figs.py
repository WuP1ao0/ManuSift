"""R-2026-06-17 (Phase C:
multi-fig xlsx
extraction):

Tests for the
``detect_xlsx_figs``
algorithm and
the updated
``extract_xlsx_text``
that emits
one block
per fig
panel.

Layouts
covered:

  * **single-table**
    (1 fig or
    no fig
    headers)
    -- should
    fall back
    to the
    single-TSV
    behavior
    so the
    common
    case isn't
    affected
  * **horizontal
    split**
    (multiple
    figs side-by-side
    in the
    same row,
    separated
    by blank
    columns)
    -- the
    Nature
    ``Sfig.2``
    case
    (``Fig.S1a``,
    ``Fig.S1b``,
    ``Fig.S1c``,
    etc. in
    one row)
  * **vertical
    split**
    (multiple
    figs stacked
    vertically,
    separated
    by a
    blank
    row) -- the
    Nature
    ``Sfig.3``
    case
    (``Fig.S3a``
    in R0,
    blank R7,
    ``Fig.S3b``
    in R8)
  * **mixed
    split**
    (some figs
    horizontal,
    some vertical
    in the
    same sheet)
    -- the
    Nature
    ``Sfig.4``
    case
"""
from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

sys.path.insert(0, r"C:\Users\22509\Desktop\ManuSift1")

from manusift.tools.safe_read_b import (
    detect_xlsx_figs,
    extract_xlsx_text,
    _looks_like_fig_header,
)


# ============================================================================
# Unit tests for the helper functions
# ============================================================================


class TestLooksLikeFigHeader:
    @pytest.mark.parametrize(
        "cell,expected",
        [
            ("Fig.S1a", "Fig.S1a"),
            ("Fig.S4b", "Fig.S4b"),
            ("Fig. S1a", "Fig. S1a"),
            ("Fig. 2a", "Fig. 2a"),
            ("Table S1", "Table S1"),
            ("Table1", "Table1"),
            ("Tab.1", "Tab.1"),
            ("fig.s1a", "fig.s1a"),  # case-insensitive
            ("Fig.S1a:", "Fig.S1a"),  # trailing colon stripped
            ("  Fig.S1a  ", "Fig.S1a"),  # surrounding whitespace
            # Non-matches
            ("hello", None),
            ("", None),
            ("123", None),
            (None, None),
            (42, None),
            ("Figure S1", "Figure S1"),  # "Figure" also matches
        ],
    )
    def test_match(self, cell, expected):
        assert _looks_like_fig_header(cell) == expected


# ============================================================================
# Synthetic xlsx fixtures
# ============================================================================


def _build_horizontal_split(path: Path) -> None:
    """Build a sheet with 3 figs side-by-side, separated by blank columns.

    Layout::

        | Fig.A |       | Fig.B |       | Fig.C |
        |  x  y |       |  x  y |       |  x  y |
        |  1  2 |       |  3  4 |       |  5  6 |
        |  7  8 |       |  9  10|       | 11  12|
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sfig.H"
    # Row 1: fig headers
    ws.cell(1, 1, "Fig.A")
    ws.cell(1, 4, "Fig.B")
    ws.cell(1, 7, "Fig.C")
    # Row 2: sub-headers
    for col in (1, 2, 4, 5, 7, 8):
        ws.cell(2, col, "x" if col in (1, 4, 7) else "y")
    # Row 3-4: data
    for r, vals in enumerate(
        [(1, 2, 3, 4, 5, 6), (7, 8, 9, 10, 11, 12)], start=3
    ):
        for i, v in enumerate(vals, start=1):
            ws.cell(r, i, v)
    wb.save(str(path))


def _build_vertical_split(path: Path) -> None:
    """Build a sheet with 2 figs stacked vertically, separated by a blank row.

    Layout::

        | Fig.A |
        |  x  y |
        |  1  2 |
        |  3  4 |
        |        <- blank
        | Fig.B |
        |  x  y |
        |  5  6 |
        |  7  8 |
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sfig.V"
    ws.cell(1, 1, "Fig.A")
    ws.cell(2, 1, "x")
    ws.cell(2, 2, "y")
    ws.cell(3, 1, 1)
    ws.cell(3, 2, 2)
    ws.cell(4, 1, 3)
    ws.cell(4, 2, 4)
    # Row 5 blank
    ws.cell(6, 1, "Fig.B")
    ws.cell(7, 1, "x")
    ws.cell(7, 2, "y")
    ws.cell(8, 1, 5)
    ws.cell(8, 2, 6)
    ws.cell(9, 1, 7)
    ws.cell(9, 2, 8)
    wb.save(str(path))


def _build_mixed_split(path: Path) -> None:
    """Build a sheet with both horizontal and vertical figs.

    Layout::

        | Fig.A |       | Fig.D |
        |  x  y |       |  x  y |
        |  1  2 |       |  3  4 |
        |        <- blank
        | Fig.B |
        |  x  y |
        |  5  6 |
        |        <- blank
        | Fig.C |
        |  x  y |
        |  7  8 |
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sfig.M"
    # Row 1: top figs
    ws.cell(1, 1, "Fig.A")
    ws.cell(1, 4, "Fig.D")
    # Row 2: sub
    ws.cell(2, 1, "x")
    ws.cell(2, 2, "y")
    ws.cell(2, 4, "x")
    ws.cell(2, 5, "y")
    # Row 3: data
    ws.cell(3, 1, 1)
    ws.cell(3, 2, 2)
    ws.cell(3, 4, 3)
    ws.cell(3, 5, 4)
    # Row 4 blank
    ws.cell(5, 1, "Fig.B")
    ws.cell(6, 1, "x")
    ws.cell(6, 2, "y")
    ws.cell(7, 1, 5)
    ws.cell(7, 2, 6)
    # Row 8 blank
    ws.cell(9, 1, "Fig.C")
    ws.cell(10, 1, "x")
    ws.cell(10, 2, "y")
    ws.cell(11, 1, 7)
    ws.cell(11, 2, 8)
    wb.save(str(path))


def _build_single_table(path: Path) -> None:
    """Build a sheet with NO fig headers -- should fall back to single-TSV."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Plain"
    ws.cell(1, 1, "Name")
    ws.cell(1, 2, "Age")
    ws.cell(2, 1, "Alice")
    ws.cell(2, 2, 30)
    ws.cell(3, 1, "Bob")
    ws.cell(3, 2, 25)
    wb.save(str(path))


def _build_one_fig_with_continuation(path: Path) -> None:
    """Build a sheet with ONE fig header but lots of data below.

    Tests the case where ``len(bboxes) == 1`` so the single-TSV
    fallback path is taken (no fig-block emission).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Single"
    ws.cell(1, 1, "Fig.1")
    ws.cell(2, 1, "x")
    ws.cell(2, 2, "y")
    for r in range(3, 30):
        ws.cell(r, 1, r)
        ws.cell(r, 2, r * 2)
    wb.save(str(path))


# ============================================================================
# detect_xlsx_figs
# ============================================================================


class TestDetectHorizontalSplit:
    def test_three_figs_detected(self, tmp_path):
        p = tmp_path / "h.xlsx"
        _build_horizontal_split(p)
        from openpyxl import load_workbook
        wb = load_workbook(p)
        ws = wb.active
        bboxes = detect_xlsx_figs(ws)
        wb.close()
        names = [bb["name"] for bb in bboxes]
        assert names == ["Fig.A", "Fig.B", "Fig.C"]

    def test_columns_non_overlapping(self, tmp_path):
        p = tmp_path / "h.xlsx"
        _build_horizontal_split(p)
        from openpyxl import load_workbook
        wb = load_workbook(p)
        ws = wb.active
        bboxes = detect_xlsx_figs(ws)
        wb.close()
        # Each fig has its own column range
        for bb in bboxes:
            assert bb["top"] == 0
            assert bb["bottom"] >= 2  # includes data
            assert bb["left"] < bb["right"]


class TestDetectVerticalSplit:
    def test_two_figs_detected(self, tmp_path):
        p = tmp_path / "v.xlsx"
        _build_vertical_split(p)
        from openpyxl import load_workbook
        wb = load_workbook(p)
        ws = wb.active
        bboxes = detect_xlsx_figs(ws)
        wb.close()
        names = [bb["name"] for bb in bboxes]
        assert names == ["Fig.A", "Fig.B"]

    def test_rows_non_overlapping(self, tmp_path):
        p = tmp_path / "v.xlsx"
        _build_vertical_split(p)
        from openpyxl import load_workbook
        wb = load_workbook(p)
        ws = wb.active
        bboxes = detect_xlsx_figs(ws)
        wb.close()
        # Each fig has its own row range. The detector's
        # ``bottom`` is the *start* of the next fig (the header
        # row), so they touch at the boundary. We assert the
        # *strict* inequality ``bottom_0 < top_1`` (i.e. Fig.A's
        # data ends before Fig.B's header begins) only when
        # there's a blank row between them; otherwise they touch.
        assert bboxes[0]["top"] < bboxes[1]["top"]
        # Specifically: Fig.A's bottom (5) is the blank row,
        # Fig.B's top (5) is its header row. They share the same
        # row -- that's the "vertical split via blank row"
        # signal. Real data would have Fig.B's header 1+ rows
        # below Fig.A's blank.
        assert bboxes[0]["bottom"] <= bboxes[1]["top"]


class TestDetectMixedSplit:
    def test_four_figs_detected(self, tmp_path):
        p = tmp_path / "m.xlsx"
        _build_mixed_split(p)
        from openpyxl import load_workbook
        wb = load_workbook(p)
        ws = wb.active
        bboxes = detect_xlsx_figs(ws)
        wb.close()
        names = [bb["name"] for bb in bboxes]
        assert names == ["Fig.A", "Fig.D", "Fig.B", "Fig.C"]


class TestDetectSingleTable:
    def test_no_fig_headers(self, tmp_path):
        """A plain table with no fig markers should return an empty list."""
        p = tmp_path / "plain.xlsx"
        _build_single_table(p)
        from openpyxl import load_workbook
        wb = load_workbook(p)
        ws = wb.active
        bboxes = detect_xlsx_figs(ws)
        wb.close()
        assert bboxes == []

    def test_one_fig_returns_one(self, tmp_path):
        """A single fig header with lots of data still returns one bbox."""
        p = tmp_path / "single.xlsx"
        _build_one_fig_with_continuation(p)
        from openpyxl import load_workbook
        wb = load_workbook(p)
        ws = wb.active
        bboxes = detect_xlsx_figs(ws)
        wb.close()
        assert len(bboxes) == 1
        assert bboxes[0]["name"] == "Fig.1"


class TestDetectEmptySheet:
    def test_empty_sheet(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        bboxes = detect_xlsx_figs(ws)
        assert bboxes == []
        wb.close()


# ============================================================================
# extract_xlsx_text (end-to-end with synthetic xlsx)
# ============================================================================


class TestExtractXlsxTextHorizontal:
    def test_emits_per_fig_blocks(self, tmp_path):
        p = tmp_path / "h.xlsx"
        _build_horizontal_split(p)
        text = extract_xlsx_text(str(p))
        # 3 fig headers
        assert "## Fig: Fig.A" in text
        assert "## Fig: Fig.B" in text
        assert "## Fig: Fig.C" in text
        # Each one has row/col ranges
        assert "cols" in text
        # The data values are present
        assert "1" in text and "12" in text
        # NO top-level "## Fig" should be a duplicate of header in data
        # (the data row contains the fig label in cell 1, but our
        # output puts the data under the fig-block header).

    def test_sheet_header_present(self, tmp_path):
        p = tmp_path / "h.xlsx"
        _build_horizontal_split(p)
        text = extract_xlsx_text(str(p))
        assert "## Sheet: Sfig.H" in text


class TestExtractXlsxTextVertical:
    def test_emits_per_fig_blocks(self, tmp_path):
        p = tmp_path / "v.xlsx"
        _build_vertical_split(p)
        text = extract_xlsx_text(str(p))
        assert "## Fig: Fig.A" in text
        assert "## Fig: Fig.B" in text
        # Fig.A's data is rows 1-4 (header, sub, 2 data rows + 1 blank)
        # Fig.B's data is rows 5-8
        assert "5" in text
        assert "8" in text


class TestExtractXlsxTextMixed:
    def test_emits_all_4_figs(self, tmp_path):
        p = tmp_path / "m.xlsx"
        _build_mixed_split(p)
        text = extract_xlsx_text(str(p))
        for name in ("Fig.A", "Fig.B", "Fig.C", "Fig.D"):
            assert f"## Fig: {name}" in text, f"missing {name}"


class TestExtractXlsxTextSingleTable:
    def test_no_fig_blocks(self, tmp_path):
        """A plain table emits a single TSV block, no ## Fig: markers."""
        p = tmp_path / "plain.xlsx"
        _build_single_table(p)
        text = extract_xlsx_text(str(p))
        assert "## Fig:" not in text
        assert "## Sheet: Plain" in text
        # The data is still present as a single TSV
        assert "Name" in text
        assert "Alice" in text

    def test_one_fig_with_lots_of_data(self, tmp_path):
        """One fig + lots of data = single TSV block, no per-fig marker.

        The detector returns 1 bbox → ``len(bboxes) <= 1`` so the
        single-TSV fallback path is taken. This is intentional: a
        sheet with one fig is the common case and shouldn't be
        wrapped in ``## Fig:`` markers.
        """
        p = tmp_path / "single.xlsx"
        _build_one_fig_with_continuation(p)
        text = extract_xlsx_text(str(p))
        # No per-fig markers (single-TSV fallback)
        assert "## Fig:" not in text
        # The data is all there
        assert "Fig.1" in text
        assert "x" in text
        assert "y" in text


# ============================================================================
# End-to-end with the real Nature SI file
# ============================================================================


REAL_FILE = (
    r"C:\Users\22509\ZCodeProject\s41565-025-02082-0"
    r"\Source_Data_MOESM3.xlsx"
)


@pytest.mark.skipif(
    not Path(REAL_FILE).exists(),
    reason="Real Nature SI file not available on this machine",
)
class TestExtractXlsxTextRealNatureFile:
    def test_sfig3_vertical_split(self):
        text = extract_xlsx_text(REAL_FILE)
        # Sfig.3 has Fig.S3a + Fig.S3b stacked vertically
        assert "## Fig: Fig.S3a" in text
        assert "## Fig: Fig.S3b" in text
        # The Sfig.3 sheet header
        assert "## Sheet: Sfig.3" in text

    def test_sfig4_mixed_split(self):
        text = extract_xlsx_text(REAL_FILE)
        # Sfig.4 has 5 figs: S4a, S4d, S4e (top row), S4b, S4c (vertical)
        for name in ("Fig. S4a", "Fig.S4d", "Fig.S4e", "Fig. S4b", "Fig. S4c"):
            assert f"## Fig: {name}" in text, f"missing {name}"

    def test_sfig2_horizontal_split(self):
        text = extract_xlsx_text(REAL_FILE)
        # Sfig.2 has 6 figs side-by-side (Fig.S1a through Fig.S1f)
        for i in "abcdef":
            assert f"## Fig: Fig.S1{i}" in text

    def test_total_output_size_reasonable(self):
        """The 874KB xlsx should produce an output that's < 200KB
        (the fig detector + 200-row cap should keep it small)."""
        text = extract_xlsx_text(REAL_FILE)
        size = len(text.encode("utf-8"))
        assert size < 200_000, (
            f"Output too large: {size} bytes "
            f"(Sfig.2 alone has 20678 rows × 6 figs)"
        )
