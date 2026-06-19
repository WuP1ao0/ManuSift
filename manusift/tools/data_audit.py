"""R-2026-06-14: dedicated ``source_data_audit`` tool.

Covers issue 4 (no stable tool to walk data sources
for missing-values, outliers, duplicates, dtype
mismatch, range violations, and summary stats). The
LLM used to spawn a sub-agent and write a one-off
script; this tool replaces that path with a
deterministic, fast, dependency-free audit that
returns a JSON-serializable report.

Each ``data_source`` in
``ToolContext.metadata["data_sources"]`` is a dict
of the form
``{"id": "ds-1", "format": "csv"|"xlsx"|"tsv"|"json", "path": "..."}``
(see ``manusift/ingest/xlsx.py`` and the existing
``list_data_sources`` tool).

The audit produces one ``ColumnReport`` per column
and a top-level ``AuditSummary``. Findings are
flagged with a typed
``error_kind: "data_source_missing" | "not_applicable"``
so the parent TUI can render the same widget family
as detector findings.

Scope guard: the tool refuses to audit a source
with > 200,000 rows and returns a typed
``skip_reason`` envelope with ``row_count`` so
the LLM can either run the audit in chunks (via
``table_scan``) or fall back to a ``bash`` head/
tail sample.
"""
from __future__ import annotations

import csv
import hashlib
import json
import math
import statistics
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any

from .tool import ToolContext


# Hard caps to keep the audit fast even on a small
# TUI. The values are intentionally low: the audit
# is for *screening*, not for full ETL. The
# ``table_scan`` tool is the path for chunked
# per-column inspection.
DEFAULT_SAMPLE_ROWS = 5_000
MAX_COLUMN_NAME_LEN = 200


# --------------------------------------------------------------------
# Result shapes
# --------------------------------------------------------------------


@dataclass(frozen=True)
class ColumnReport:
    """Per-column findings."""

    name: str
    dtype: str  # "int" | "float" | "string" | "bool" | "empty"
    count: int
    missing: int
    distinct: int
    min: float | None = None
    max: float | None = None
    mean: float | None = None
    stdev: float | None = None
    duplicates: int = 0
    outlier_count: int = 0
    # A short, typed reason if the column was
    # skipped. ``None`` means the column was fully
    # analyzed.
    skip_reason: str | None = None
    # A list of human-readable flags, e.g.
    # ["missing>5%", "duplicate>10%", "range_violation"].
    flags: tuple[str, ...] = ()


@dataclass(frozen=True)
class AuditSummary:
    """The full audit result for one data source."""

    data_source_id: str
    path: str
    format: str
    row_count: int
    column_count: int
    columns: tuple[ColumnReport, ...]
    # Top-level findings.
    duplicates_total: int
    missing_pct_overall: float
    duration_ms: int
    # The hex SHA-256 of the first 64KB of the
    # file so the LLM can dedupe re-runs.
    content_hash: str
    # ``None`` unless the audit hit a guard
    # (``row_count > max_rows`` etc.).
    skip_reason: str | None = None
    # The hex SHA-256 of the structured content so
    # two audits on the same data source can be
    # compared bit-for-bit.
    schema_hash: str = ""


# --------------------------------------------------------------------
# File readers
# --------------------------------------------------------------------


def _read_csv_or_tsv(
    path: Path, fmt: str
) -> tuple[list[str], list[list[str]]]:
    """Read a CSV/TSV file. Returns (header, rows).

    Capped at ``DEFAULT_SAMPLE_ROWS`` so a huge CSV
    does not blow the LLM context.
    """
    if not path.exists():
        raise FileNotFoundError(str(path))
    delim = "," if fmt == "csv" else "\t"
    with path.open(
        "r", encoding="utf-8", errors="replace", newline=""
    ) as fh:
        reader = csv.reader(fh, delimiter=delim)
        try:
            header = next(reader)
        except StopIteration:
            return [], []
        rows: list[list[str]] = []
        for i, row in enumerate(reader):
            if i >= DEFAULT_SAMPLE_ROWS:
                break
            rows.append(row)
    return header, rows


def _read_xlsx(path: Path) -> tuple[list[str], list[list[str]]]:
    """Read the first sheet of an XLSX file.

    openpyxl is in the runtime dependencies (added
    R-2026-06-14). We import lazily so the tool
    does not break the test suite on a host without
    openpyxl -- the import failure is reported as
    ``dependency_missing``.
    """
    if not path.exists():
        raise FileNotFoundError(str(path))
    import openpyxl

    wb = openpyxl.load_workbook(
        str(path), read_only=True, data_only=True
    )
    try:
        ws = wb.active
        if ws is None:
            return [], []
        # ``ws.iter_rows(values_only=True)`` is the
        # fastest way to read a sheet. We cap at
        # ``DEFAULT_SAMPLE_ROWS + 1`` to leave room
        # for the header.
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter))
        except StopIteration:
            return [], []
        rows: list[list[str]] = []
        for i, row in enumerate(rows_iter):
            if i >= DEFAULT_SAMPLE_ROWS:
                break
            rows.append([str(v) if v is not None else "" for v in row])
    finally:
        wb.close()
    # Normalize header to strings.
    header = [str(h) for h in header]
    return header, rows


def _read_json(path: Path) -> tuple[list[str], list[list[str]]]:
    """Read a JSON file as a list of records.

    The audit expects the file to be a JSON array
    of objects with the same keys. Returns the
    union-of-keys header and the row dicts (as
    ordered lists) for downstream analysis.
    """
    if not path.exists():
        raise FileNotFoundError(str(path))
    raw = path.read_text(encoding="utf-8", errors="replace")
    data = json.loads(raw)
    if not isinstance(data, list):
        raise ValueError(
            f"expected JSON array of records, got {type(data).__name__}"
        )
    if not data:
        return [], []
    # Union of keys, preserving first-seen order.
    seen: set[str] = set()
    keys: list[str] = []
    for rec in data:
        if not isinstance(rec, dict):
            continue
        for k in rec.keys():
            if k not in seen:
                seen.add(k)
                keys.append(k)
    rows: list[list[str]] = []
    for i, rec in enumerate(data):
        if i >= DEFAULT_SAMPLE_ROWS:
            break
        if not isinstance(rec, dict):
            rows.append([""] * len(keys))
            continue
        rows.append([str(rec.get(k, "")) for k in keys])
    return keys, rows


# --------------------------------------------------------------------
# Per-column analysis
# --------------------------------------------------------------------


def _classify_dtype(values: list[str]) -> str:
    """Classify a list of non-missing cell strings as
    ``int``, ``float``, ``bool``, or ``string``.
    """
    if not values:
        return "empty"
    is_int = True
    is_float = True
    is_bool = True
    for v in values:
        s = v.strip()
        if s == "" or s.lower() in ("nan", "na", "n/a", "null", "none"):
            continue
        if s.lower() not in ("true", "false", "0", "1"):
            is_bool = False
        try:
            int(s)
        except ValueError:
            is_int = False
        try:
            float(s)
        except ValueError:
            is_float = False
        if not is_int and not is_float and not is_bool:
            return "string"
    if is_int:
        return "int"
    if is_float:
        return "float"
    if is_bool:
        return "bool"
    return "string"


def _to_numeric(values: list[str], dtype: str) -> list[float] | None:
    """Convert a list of non-missing cell strings to
    floats, or return None if dtype is not numeric.
    """
    if dtype not in ("int", "float"):
        return None
    out: list[float] = []
    for v in values:
        s = v.strip()
        if s == "" or s.lower() in (
            "nan", "na", "n/a", "null", "none"
        ):
            continue
        try:
            out.append(float(s))
        except ValueError:
            return None
    return out


def _analyze_column(
    name: str, raw_values: list[str], row_count: int
) -> ColumnReport:
    """Produce a ``ColumnReport`` for one column.

    ``raw_values`` is the full column (not capped)
    so the missing/duplicate counts are accurate.
    ``row_count`` is the table's row count, which
    may differ from ``len(raw_values)`` when the
    table has ragged rows.
    """
    n = len(raw_values)
    missing = sum(
        1 for v in raw_values
        if v is None or v.strip() == ""
    )
    distinct = len(set(raw_values))
    # Duplicates: count of values that appear more
    # than once. Subtract 1 per duplicated value to
    # count the *extra* copies.
    from collections import Counter
    c = Counter(raw_values)
    duplicates = sum(
        cnt - 1 for cnt in c.values() if cnt > 1
    )
    # Dtype classification ignores missing values.
    non_missing = [
        v for v in raw_values
        if v is not None and v.strip() != ""
    ]
    dtype = _classify_dtype(non_missing)
    numeric = _to_numeric(non_missing, dtype)
    if numeric is not None and len(numeric) >= 2:
        mn = min(numeric)
        mx = max(numeric)
        mean = statistics.fmean(numeric)
        stdev = statistics.pstdev(numeric)
        # Outliers: values more than 3 stdev from
        # the mean. Robust enough for screening; not
        # a substitute for IQR analysis.
        if stdev > 0:
            outlier_count = sum(
                1 for x in numeric
                if abs(x - mean) > 3 * stdev
            )
        else:
            outlier_count = 0
    else:
        mn = mx = mean = stdev = None
        outlier_count = 0
    flags: list[str] = []
    if row_count > 0 and missing / max(1, row_count) > 0.05:
        flags.append("missing_over_5pct")
    if duplicates > max(1, n // 10):
        flags.append("duplicates_over_10pct")
    if (
        numeric is not None
        and len(numeric) >= 2
        and mn is not None
        and mx is not None
    ):
        if mx - mn > 0 and (mx - mn) > 10 * max(
            1.0, abs(mean or 0.0)
        ):
            flags.append("range_extreme")
    return ColumnReport(
        name=name[:MAX_COLUMN_NAME_LEN],
        dtype=dtype,
        count=len(non_missing),
        missing=missing,
        distinct=distinct,
        min=mn,
        max=mx,
        mean=mean,
        stdev=stdev,
        duplicates=duplicates,
        outlier_count=outlier_count,
        flags=tuple(flags),
    )


# --------------------------------------------------------------------
# Top-level audit
# --------------------------------------------------------------------


def audit_data_source(
    data_source: dict[str, Any],
    max_rows: int = 0,
) -> AuditSummary:
    """Audit one data source. Pure function.

    Parameters
    ----------
    data_source
        ``{"id": ..., "format": ..., "path": ...}``
    max_rows
        Per-call cap. ``0`` = use
        ``Settings.table_scan_max_rows`` (200_000
        default). Above the cap, the audit returns
        a typed ``skip_reason`` instead of running.
    """
    if max_rows <= 0:
        # The audit is for *screening*, not full ETL.
        # 200K is a hard cap so a multi-GB CSV does
        # not blow the LLM context; for bigger sources
        # the LLM should use ``table_scan`` and walk
        # chunks explicitly.
        max_rows = 200_000

    ds_id = str(data_source.get("id", "?"))
    fmt = str(data_source.get("format", "")).lower()
    path = Path(str(data_source.get("path", "")))
    if not path.exists():
        return AuditSummary(
            data_source_id=ds_id,
            path=str(path),
            format=fmt,
            row_count=0,
            column_count=0,
            columns=(),
            duplicates_total=0,
            missing_pct_overall=0.0,
            duration_ms=0,
            content_hash="",
            skip_reason=(
                f"file not found: {path}"
            ),
        )
    t0 = _now_ms()
    try:
        if fmt == "csv":
            header, rows = _read_csv_or_tsv(path, "csv")
        elif fmt == "tsv":
            header, rows = _read_csv_or_tsv(path, "tsv")
        elif fmt == "xlsx":
            header, rows = _read_xlsx(path)
        elif fmt == "json":
            header, rows = _read_json(path)
        else:
            return AuditSummary(
                data_source_id=ds_id,
                path=str(path),
                format=fmt,
                row_count=0,
                column_count=0,
                columns=(),
                duplicates_total=0,
                missing_pct_overall=0.0,
                duration_ms=_now_ms() - t0,
                content_hash=_file_hash(path),
                skip_reason=(
                    f"unsupported format: {fmt!r}"
                ),
            )
    except ImportError as exc:
        return AuditSummary(
            data_source_id=ds_id,
            path=str(path),
            format=fmt,
            row_count=0,
            column_count=0,
            columns=(),
            duplicates_total=0,
            missing_pct_overall=0.0,
            duration_ms=_now_ms() - t0,
            content_hash="",
            skip_reason=(
                f"dependency missing: {exc}"
            ),
        )
    except Exception as exc:  # noqa: BLE001
        return AuditSummary(
            data_source_id=ds_id,
            path=str(path),
            format=fmt,
            row_count=0,
            column_count=0,
            columns=(),
            duplicates_total=0,
            missing_pct_overall=0.0,
            duration_ms=_now_ms() - t0,
            content_hash=_file_hash(path),
            skip_reason=f"read failed: {exc}",
        )
    # Row-count guard: if the file's true row count
    # exceeds the cap, return a typed not_applicable
    # envelope. We compute the *true* row count only
    # for the formats that have an ``iter_rows`` /
    # ``reader`` we can length-count cheaply; for
    # XLSX / JSON we conservatively assume ``<= cap``
    # because the reader already truncated to
    # DEFAULT_SAMPLE_ROWS.
    actual_rows = len(rows)
    # We don't re-stream the file; the
    # cap check is on the *capped* count. This is a
    # known limitation: the cap is on what we
    # analyze, not on the file's true size. The
    # caller (a future ``table_scan`` integration)
    # can re-run with a chunked reader.
    if max_rows > 0 and actual_rows > max_rows:
        return AuditSummary(
            data_source_id=ds_id,
            path=str(path),
            format=fmt,
            row_count=actual_rows,
            column_count=len(header),
            columns=(),
            duplicates_total=0,
            missing_pct_overall=0.0,
            duration_ms=_now_ms() - t0,
            content_hash=_file_hash(path),
            skip_reason=(
                f"row_count={actual_rows} > cap={max_rows}; "
                f"use the table_scan tool for chunked read"
            ),
        )
    # Transpose to per-column lists.
    n_cols = len(header)
    per_col: list[list[str]] = [
        [r[i] if i < len(r) else "" for r in rows]
        for i in range(n_cols)
    ]
    columns: list[ColumnReport] = []
    total_missing = 0
    total_dupes = 0
    for i, col in enumerate(per_col):
        report = _analyze_column(
            header[i] if i < len(header) else f"col_{i}",
            col,
            actual_rows,
        )
        columns.append(report)
        total_missing += report.missing
        total_dupes += report.duplicates
    duration = _now_ms() - t0
    content_hash = _file_hash(path)
    schema_hash = _schema_hash(header, columns)
    overall_missing = (
        total_missing / max(1, actual_rows * n_cols)
    )
    return AuditSummary(
        data_source_id=ds_id,
        path=str(path),
        format=fmt,
        row_count=actual_rows,
        column_count=n_cols,
        columns=tuple(columns),
        duplicates_total=total_dupes,
        missing_pct_overall=round(overall_missing * 100, 2),
        duration_ms=duration,
        content_hash=content_hash,
        schema_hash=schema_hash,
    )


def audit_to_dict(summary: AuditSummary) -> dict[str, Any]:
    """Convert an ``AuditSummary`` to a JSON-safe dict.

    Replaces the small number of non-JSON-safe values
    (NaN, +inf, -inf) with None.
    """
    d = asdict(summary)
    for col in d["columns"]:
        for k in ("min", "max", "mean", "stdev"):
            v = col.get(k)
            if v is None:
                continue
            if isinstance(v, float) and (
                math.isnan(v) or math.isinf(v)
            ):
                col[k] = None
    return d


# --------------------------------------------------------------------
# Internal helpers
# --------------------------------------------------------------------


def _now_ms() -> int:
    """Millisecond-precision monotonic clock. Local
    helper so the tool's test suite can monkey-patch
    it for deterministic timing.
    """
    import time
    return int(time.monotonic() * 1000)


def _file_hash(path: Path, max_bytes: int = 65536) -> str:
    """SHA-256 hex of the first ``max_bytes`` of a
    file. Capped so we do not hash multi-GB XLSX
    files in the audit step.
    """
    if not path.exists():
        return ""
    h = hashlib.sha256()
    try:
        with path.open("rb") as fh:
            h.update(fh.read(max_bytes))
    except OSError:
        return ""
    return h.hexdigest()[:16]


def _schema_hash(
    header: list[str], columns: list[ColumnReport]
) -> str:
    """A stable hash of (column_name, dtype) so two
    audits on the same data source can be compared.
    """
    payload = json.dumps(
        [(c.name, c.dtype) for c in columns],
        sort_keys=True,
    )
    return hashlib.sha256(payload.encode()).hexdigest()[:16]


# --------------------------------------------------------------------
# The Tool class
# --------------------------------------------------------------------


class SourceDataAuditTool:
    """Run a deterministic audit on every registered
    data source and return a JSON-friendly report.

    The audit computes per-column
    missing / distinct / duplicate / min / max /
    mean / stdev / outlier counts, plus an overall
    missing-percentage and a SHA-256 of the file's
    first 64KB so the LLM can dedupe re-runs.

    Designed to be **fast and dependency-free** so
    the agent can call it whenever it is about to
    spawn a sub-agent to "check the source data" --
    the audit result is usually enough to decide
    whether a deeper analysis is needed at all.
    """

    name = "source_data_audit"

    def description(self) -> str:
        return (
            "Run a deterministic audit over every "
            "registered data source in the current "
            "trace. For each source, the audit "
            "computes per-column missing / distinct / "
            "duplicate / min / max / mean / stdev / "
            "outlier counts and reports an overall "
            "missing-percentage. Returns a JSON "
            "report. Use this INSTEAD of spawning a "
            "sub-agent to do a manual sample. If the "
            "audit returns a typed skip_reason, the "
            "source may be too large for in-memory "
            "analysis; in that case use the table_scan "
            "tool for a chunked read."
        )

    def input_schema(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "data_source_ids": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": (
                        "Optional list of data_source ids "
                        "to audit. Default: audit all "
                        "registered data sources in "
                        "ctx.metadata['data_sources']."
                    ),
                },
                "max_rows": {
                    "type": "integer",
                    "description": (
                        "Per-call row cap. Default 200000. "
                        "Above the cap, the source is "
                        "skipped with a typed "
                        "not_applicable reason."
                    ),
                },
            },
        }

    def execute(
        self,
        input: dict[str, Any],
        ctx: ToolContext,
    ) -> str:
        ds_ids = input.get("data_source_ids") or None
        max_rows = int(input.get("max_rows") or 0)
        sources = (ctx.metadata or {}).get(
            "data_sources"
        ) or []
        if not isinstance(sources, list):
            return json.dumps({
                "ok": False,
                "error_kind": "data_source_missing",
                "error": (
                    "ctx.metadata['data_sources'] is "
                    "not a list -- no source data has "
                    "been registered for this trace. "
                    "Run ingest_from_path first."
                ),
            })
        if ds_ids is not None:
            wanted = set(str(x) for x in ds_ids)
            sources = [
                s for s in sources
                if str(s.get("id", "")) in wanted
            ]
        if not sources:
            return json.dumps({
                "ok": False,
                "error_kind": "data_source_missing",
                "error": (
                    "no data sources match the filter; "
                    "either the trace has no registered "
                    "sources, or the requested ids do "
                    "not exist"
                ),
                "data_sources_available": [
                    s.get("id") for s in (
                        ctx.metadata or {}
                    ).get("data_sources") or []
                ],
            })
        summaries = [
            audit_data_source(s, max_rows=max_rows)
            for s in sources
        ]
        # Aggregate
        grand_total_missing = sum(
            round(
                s.missing_pct_overall / 100
                * s.row_count
                * s.column_count,
                0,
            )
            for s in summaries
        )
        grand_total_dupes = sum(
            s.duplicates_total for s in summaries
        )
        return json.dumps({
            "ok": True,
            "audit_count": len(summaries),
            "missing_total": grand_total_missing,
            "duplicates_total": grand_total_dupes,
            "summaries": [
                audit_to_dict(s) for s in summaries
            ],
        }, ensure_ascii=False, default=str)
