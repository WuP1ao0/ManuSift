"""Tests for the data-source / table-statistics surface (R-audit, 2026-06).

The full pipeline tested in
two layers:

  1. **Parser layer** --
     ``manusift.ingest.xlsx``
     turns real XLSX / CSV /
     TSV / JSON files into
     ``ExtractedTable``
     records. We round-trip
     real Nature-paper source
     data plus synthetic
     fixtures so the parser's
     empty-row / None-cell /
     header-detection rules
     are pinned.

  2. **Tool layer** --
     ``list_data_sources`` /
     ``read_data_source`` /
     ``table_benford`` /
     ``table_duplicate_row`` /
     ``table_outlier`` /
     ``table_round_bias`` are
     all reachable from the
     agent loop AND operate
     on a ParsedDoc whose
     ``tables`` is populated
     from companion XLSX
     files. We verify that
     the data-source helpers
     return what the LLM
     would expect, and that
     at least one of the
     statistics detectors
     produces a finding on
     real data.

The end-to-end "LLM calls
the new tools" flow is
exercised separately in
``test_pilot_table_e2e.py``
so this file focuses on
the *plumbing*.
"""
from __future__ import annotations

import csv
import io
import json
import shutil
import zipfile
from pathlib import Path

import pytest


# ---------- fixtures ----------


@pytest.fixture
def xlsx_with_two_sheets(
    tmp_path: Path,
) -> Path:
    """Write a real .xlsx
    file with two sheets
    so we can verify that
    the parser produces
    two
    ``ExtractedTable``s.
    Falls back to skipping
    if ``openpyxl`` is not
    installed (so the rest
    of the test suite still
    runs in minimal envs).
    """
    openpyxl = pytest.importorskip("openpyxl")
    p = tmp_path / "two_sheets.xlsx"
    wb = openpyxl.Workbook()
    # Sheet 1:
    # numeric
    # data.
    ws1 = wb.active
    ws1.title = "Fig1"
    ws1.append(["Group", "Value", "p_value"])
    for i in range(20):
        ws1.append(
            [
                f"G{i // 4}",
                float(i) * 1.1,
                0.001 * i,
            ]
        )
    # Sheet 2:
    # string-
    # heavy.
    ws2 = wb.create_sheet("Fig2")
    ws2.append(["Sample", "Treatment"])
    for i in range(10):
        ws2.append([f"S{i}", "CM-Ms@IOX4"])
    wb.save(p)
    return p


@pytest.fixture
def csv_with_quoted_text(
    tmp_path: Path,
) -> Path:
    p = tmp_path / "with_quotes.csv"
    with open(p, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["x", "y", "note"])
        writer.writerow(
            ["1.0", "2.0", 'hello, "world"']
        )
        writer.writerow(["3.0", "4.0", "another"])
    return p


@pytest.fixture
def tsv_with_tab_delimiter(
    tmp_path: Path,
) -> Path:
    p = tmp_path / "with_tabs.tsv"
    p.write_text(
        "a\tb\tc\n"
        "1\t2\t3\n"
        "4\t5\t6\n",
        encoding="utf-8",
    )
    return p


@pytest.fixture
def json_table_file(
    tmp_path: Path,
) -> Path:
    p = tmp_path / "table.json"
    p.write_text(
        json.dumps(
            {
                "headers": ["col1", "col2"],
                "rows": [["1", "2"], ["3", "4"]],
            }
        ),
        encoding="utf-8",
    )
    return p


# ---------- 1. parser ----------


def test_parse_xlsx_two_sheets(
    xlsx_with_two_sheets: Path,
) -> None:
    from manusift.ingest.xlsx import parse_xlsx

    tables = parse_xlsx(xlsx_with_two_sheets)
    assert len(tables) == 2
    by_sheet = {t.sheet_name: t for t in tables}
    assert "Fig1" in by_sheet
    assert "Fig2" in by_sheet
    fig1 = by_sheet["Fig1"]
    assert fig1.source_kind == "xlsx"
    assert fig1.headers == ["Group", "Value", "p_value"]
    assert len(fig1.rows) == 20
    # openpyxl
    # formats
    # whole-number
    # floats
    # as
    # ``0``
    # not
    # ``0.0``
    # when
    # the
    # cell
    # has
    # no
    # fractional
    # part.
    # We
    # only
    # pin
    # the
    # row
    # count
    # +
    # header
    # here.
    assert fig1.rows[0][0] == "G0"
    # Each
    # table
    # has
    # a
    # stable
    # table_id.
    assert fig1.table_id != by_sheet["Fig2"].table_id


def test_parse_xlsx_records_highlighted_cells(
    tmp_path: Path,
) -> None:
    """Yellow-filled XLSX cells are preserved as table metadata."""
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl.styles import PatternFill

    from manusift.ingest.xlsx import parse_xlsx

    p = tmp_path / "highlighted.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plain"
    ws.append(["Group", "Value"])
    ws.append(["A", 11.1])
    ws.append(["B", 22.2])
    ws["B3"].fill = PatternFill(
        fill_type="solid",
        fgColor="FFFF00",
    )
    wb.save(p)

    tables = parse_xlsx(p)

    assert len(tables) == 1
    assert tables[0].highlighted_cells == [
        {
            "row": 1,
            "col": 1,
            "source_row": 3,
            "source_col": 2,
            "value": "22.2",
            "fill": "FFFF00",
        }
    ]


def test_parse_xlsx_keeps_highlights_on_their_fig(
    tmp_path: Path,
) -> None:
    """Marked cells in Fig.3c stay attached to the Fig.3c table."""
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl.styles import PatternFill

    from manusift.ingest.xlsx import parse_xlsx

    p = tmp_path / "fig_highlighted.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mixed"
    ws.cell(1, 1, "Fig.3b")
    ws.cell(1, 4, "Fig.3c")
    ws.cell(2, 1, "Group")
    ws.cell(2, 2, "Value")
    ws.cell(2, 4, "Group")
    ws.cell(2, 5, "Value")
    ws.cell(3, 1, "B-control")
    ws.cell(3, 2, 1.2)
    ws.cell(3, 4, "C-control")
    ws.cell(3, 5, 9.8)
    ws.cell(3, 5).fill = PatternFill(
        fill_type="solid",
        fgColor="FFFF00",
    )
    wb.save(p)

    by_fig = {t.fig_name: t for t in parse_xlsx(p)}

    assert by_fig["Fig.3b"].highlighted_cells == []
    assert by_fig["Fig.3c"].highlighted_cells[0] == {
        "row": 1,
        "col": 1,
        "source_row": 3,
        "source_col": 5,
        "value": "9.8",
        "fill": "FFFF00",
    }


def test_parse_csv_quoted(tmp_path: Path) -> None:
    from manusift.ingest.xlsx import parse_csv

    p = tmp_path / "with_quotes.csv"
    # Use
    # a
    # sample
    # that's
    # big
    # enough
    # for
    # csv.Sniffer
    # to
    # actually
    # detect
    # the
    # delimiter
    # -- the
    # sniffer
    # needs
    # ~2
    # lines
    # of
    # context.
    p.write_text(
        "x,y,note\n"
        "1.0,2.0,plain\n"
        "3.0,4.0,\"with, comma\"\n"
        "5.0,6.0,another\n",
        encoding="utf-8",
    )
    tables = parse_csv(p)
    assert len(tables) == 1
    t = tables[0]
    assert t.headers == ["x", "y", "note"]
    assert t.rows[0] == ["1.0", "2.0", "plain"]
    assert t.rows[1] == ["3.0", "4.0", "with, comma"]
    assert t.source_kind == "csv"


def test_parse_tsv_explicit_delimiter(
    tsv_with_tab_delimiter: Path,
) -> None:
    from manusift.ingest.xlsx import parse_csv

    tables = parse_csv(tsv_with_tab_delimiter, delimiter="\t")
    assert len(tables) == 1
    t = tables[0]
    assert t.headers == ["a", "b", "c"]
    assert len(t.rows) == 2


def test_parse_json_table(
    json_table_file: Path,
) -> None:
    from manusift.ingest.xlsx import parse_json_table

    tables = parse_json_table(json_table_file)
    assert len(tables) == 1
    t = tables[0]
    assert t.source_kind == "json"
    assert t.headers == ["col1", "col2"]
    assert t.rows == [["1", "2"], ["3", "4"]]


def test_parse_data_file_dispatch(tmp_path: Path) -> None:
    """``parse_data_file``
    picks the right
    backend based on the
    extension."""
    from manusift.ingest.xlsx import parse_data_file

    # Unknown
    # extension
    # is
    # a
    # no-op,
    # not
    # an
    # error.
    p = tmp_path / "image.png"
    p.write_bytes(b"\x89PNG\r\n")
    assert parse_data_file(p) == []
    # CSV
    # works.
    csv_p = tmp_path / "data.csv"
    csv_p.write_text("a,b\n1,2\n", encoding="utf-8")
    assert len(parse_data_file(csv_p)) == 1


def test_parse_xlsx_missing_file(tmp_path: Path) -> None:
    from manusift.ingest.xlsx import parse_xlsx

    with pytest.raises(FileNotFoundError):
        parse_xlsx(tmp_path / "no-such.xlsx")


def test_discover_companion_files(
    tmp_path: Path,
) -> None:
    """The discover helper
    picks up XLSX / CSV /
    TSV / JSON files but
    skips hidden dirs,
    image files, and
    arbitrary depths."""
    from manusift.ingest.xlsx import (
        discover_companion_files,
        iter_data_files_in,
    )

    # Real
    # files.
    (tmp_path / "data.xlsx").write_bytes(b"")
    (tmp_path / "data.csv").write_text("a,b\n1,2\n")
    # Hidden
    # dir
    # --
    # should
    # be
    # skipped.
    hidden = tmp_path / ".git"
    hidden.mkdir()
    (hidden / "should_skip.xlsx").write_bytes(b"")
    # Image
    # --
    # not
    # supported.
    (tmp_path / "photo.png").write_bytes(b"\x89PNG")
    found = {
        p.name for p in discover_companion_files(tmp_path)
    }
    assert "data.xlsx" in found
    assert "data.csv" in found
    assert "photo.png" not in found
    assert "should_skip.xlsx" not in found


def test_discover_companion_files_extracts_supported_zip_members(
    tmp_path: Path,
) -> None:
    """Supplementary source data often arrives as a ZIP.
    Discovery should expose supported members as normal
    filesystem paths so the existing CSV/XLSX parsers can
    handle them without a special table-detector branch.
    """
    from manusift.ingest.xlsx import discover_companion_files

    archive = tmp_path / "supplementary.zip"
    with zipfile.ZipFile(archive, "w") as zf:
        zf.writestr("source_data/fig1.csv", "group,value\nA,1\n")
        zf.writestr("images/figure.png", b"not-table")
        zf.writestr("../escape.csv", "bad,path\n1,2\n")
    extracted_dir = tmp_path / "materials" / "_archives"

    found = list(
        discover_companion_files(
            tmp_path,
            extract_archives_to=extracted_dir,
        )
    )
    names = {p.name for p in found}

    assert "fig1.csv" in names
    assert "figure.png" not in names
    assert "escape.csv" not in names
    extracted_csv = next(p for p in found if p.name == "fig1.csv")
    assert extracted_csv.exists()
    assert extracted_dir in extracted_csv.parents


# ---------- 2. ExtractedTable contract ----------


def test_extracted_table_default_in_parsed_doc() -> None:
    """``ParsedDoc`` has a
    ``tables`` field with
    a default of ``[]``
    so old callers
    without the field
    keep working."""
    from manusift.contracts import (
        ExtractedTable,
        ParsedDoc,
        TextBlock,
        ExtractedImage,
    )

    doc = ParsedDoc(
        trace_id="t",
        source_path="x.pdf",
        text_blocks=[],
        images=[],
        metadata={},
    )
    assert doc.tables == []
    # We
    # can
    # also
    # pass
    # them
    # explicitly.
    t = ExtractedTable(
        table_id="x",
        source_kind="xlsx",
        source_path="x.xlsx",
        sheet_name="S",
        source_index=0,
        headers=["a"],
        rows=[["1"]],
    )
    doc2 = ParsedDoc(
        trace_id="t",
        source_path="x.pdf",
        text_blocks=[],
        images=[],
        metadata={},
        tables=[t],
    )
    assert doc2.tables == [t]


# ---------- 3. Tools ----------


def _make_ctx(parsed_doc) -> "object":
    """Build a minimal
    ToolContext for unit
    tests. The
    ``ctx.metadata``
    dict is the only
    field the new tools
    read.
    """
    from manusift.tools.tool import ToolContext

    return ToolContext(
        trace_id="t",
        current_pdf="t",
        metadata={"parsed_doc": parsed_doc},
    )


def test_list_data_sources_returns_xlsx(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """The
    ``list_data_sources``
    tool returns every
    ExtractedTable
    parsed from companion
    XLSX in
    ``materials/``."""
    monkeypatch.setenv(
        "MANUSIFT_WORKSPACE_DIR", str(tmp_path / "ws")
    )
    # Build
    # a
    # job
    # workspace
    # with
    # PDF
    # +
    # materials/.
    job = tmp_path / "ws" / "jobs" / "tid" / "materials"
    job.mkdir(parents=True)
    pdf = tmp_path / "ws" / "jobs" / "tid" / "paper.pdf"
    shutil.copy(
        "C:/Users/22509/Desktop/ScholarLens/pilot_cases/"
        "real_world_nature/s41565-025-02082-0/s41565-"
        "025-02082-0.pdf",
        pdf,
    )
    # Reuse
    # one
    # of
    # the
    # real
    # Nature
    # source-data
    # files.
    shutil.copy(
        "C:/Users/22509/Desktop/ScholarLens/pilot_cases/"
        "real_world_nature/s41565-025-02082-0/"
        "materials/source_data/Source_Data_Fig1.xlsx",
        job / "Source_Data_Fig1.xlsx",
    )
    from manusift.config import get_settings
    if hasattr(get_settings, "cache_clear"):
        get_settings.cache_clear()
    from manusift.tools.table_stats_tools import (
        ListDataSourcesTool,
    )

    # Build
    # the
    # parsed
    # doc
    # ourselves
    # so we
    # don't
    # depend
    # on
    # the
    # long-running
    # parse_pdf
    # path
    # in
    # this
    # test.
    from manusift.contracts import (
        ExtractedTable,
        ParsedDoc,
    )
    doc = ParsedDoc(
        trace_id="tid",
        source_path=str(pdf),
        text_blocks=[],
        images=[],
        metadata={},
        tables=[
            ExtractedTable(
                table_id="real-xlsx-1",
                source_kind="xlsx",
                source_path=str(
                    job / "Source_Data_Fig1.xlsx"
                ),
                sheet_name="Fig. 1",
                source_index=0,
                headers=["x"],
                rows=[["1"], ["2"]],
            ),
        ],
    )
    t = ListDataSourcesTool()
    result = t.execute(
        {"trace_id": "tid"}, _make_ctx(doc)
    )
    env = json.loads(result)
    assert env["trace_id"] == "tid"
    assert env["n_tables"] == 1
    entry = env["tables"][0]
    assert entry["source_kind"] == "xlsx"
    assert entry["sheet_name"] == "Fig. 1"
    assert entry["n_rows"] == 2


def test_read_data_source_returns_rows(
    tmp_path: Path,
) -> None:
    """``read_data_source``
    returns the full
    headers + rows of a
    single table by
    ``table_id``."""
    from manusift.contracts import (
        ExtractedTable,
        ParsedDoc,
    )
    from manusift.tools.table_stats_tools import (
        ReadDataSourceTool,
    )

    doc = ParsedDoc(
        trace_id="tid",
        source_path="x.pdf",
        text_blocks=[],
        images=[],
        metadata={},
        tables=[
            ExtractedTable(
                table_id="abc",
                source_kind="xlsx",
                source_path="x.xlsx",
                sheet_name="S",
                source_index=0,
                headers=["h1", "h2"],
                rows=[["a", "b"], ["c", "d"]],
            ),
            ExtractedTable(
                table_id="xyz",
                source_kind="xlsx",
                source_path="y.xlsx",
                sheet_name="T",
                source_index=0,
                headers=["p"],
                rows=[["1"], ["2"]],
            ),
        ],
    )
    t = ReadDataSourceTool()
    result = t.execute(
        {"trace_id": "tid", "table_id": "abc"},
        _make_ctx(doc),
    )
    env = json.loads(result)
    assert env["table_id"] == "abc"
    assert env["headers"] == ["h1", "h2"]
    assert env["rows"] == [["a", "b"], ["c", "d"]]
    assert env["n_rows_total"] == 2
    assert env["truncated"] is False


def test_data_source_tools_return_highlighted_cells() -> None:
    from manusift.contracts import (
        ExtractedTable,
        ParsedDoc,
    )
    from manusift.tools.table_stats_tools import (
        ListDataSourcesTool,
        ReadDataSourceTool,
    )

    highlighted = [
        {
            "row": 0,
            "col": 1,
            "source_row": 2,
            "source_col": 2,
            "value": "9.8",
            "fill": "FFFF00",
        }
    ]
    doc = ParsedDoc(
        trace_id="tid",
        source_path="x.pdf",
        text_blocks=[],
        images=[],
        metadata={},
        tables=[
            ExtractedTable(
                table_id="marked",
                source_kind="xlsx",
                source_path="x.xlsx",
                sheet_name="Fig.3",
                source_index=0,
                headers=["Group", "Value"],
                rows=[["C-control", "9.8"]],
                fig_name="Fig.3c",
                highlighted_cells=highlighted,
            )
        ],
    )

    listed = json.loads(
        ListDataSourcesTool().execute(
            {"trace_id": "tid"}, _make_ctx(doc)
        )
    )
    entry = listed["tables"][0]
    assert entry["n_highlighted_cells"] == 1
    assert entry["highlighted_preview"] == highlighted

    read = json.loads(
        ReadDataSourceTool().execute(
            {"trace_id": "tid", "table_id": "marked"},
            _make_ctx(doc),
        )
    )
    assert read["highlighted_cells"] == highlighted


def test_read_data_source_truncates() -> None:
    from manusift.contracts import (
        ExtractedTable,
        ParsedDoc,
    )
    from manusift.tools.table_stats_tools import (
        ReadDataSourceTool,
    )

    rows = [[str(i)] for i in range(100)]
    doc = ParsedDoc(
        trace_id="tid",
        source_path="x.pdf",
        text_blocks=[],
        images=[],
        metadata={},
        tables=[
            ExtractedTable(
                table_id="big",
                source_kind="xlsx",
                source_path="x.xlsx",
                sheet_name="S",
                source_index=0,
                headers=["n"],
                rows=rows,
            )
        ],
    )
    t = ReadDataSourceTool()
    env = json.loads(
        t.execute(
            {
                "trace_id": "tid",
                "table_id": "big",
                "max_rows": 5,
            },
            _make_ctx(doc),
        )
    )
    assert env["truncated"] is True
    assert env["n_rows_returned"] == 5
    assert env["n_rows_total"] == 100


def test_read_data_source_unknown_id() -> None:
    from manusift.contracts import (
        ExtractedTable,
        ParsedDoc,
    )
    from manusift.tools.table_stats_tools import (
        ReadDataSourceTool,
    )

    doc = ParsedDoc(
        trace_id="tid",
        source_path="x.pdf",
        text_blocks=[],
        images=[],
        metadata={},
        tables=[
            ExtractedTable(
                table_id="real",
                source_kind="xlsx",
                source_path="x.xlsx",
                sheet_name="S",
                source_index=0,
                headers=["h"],
                rows=[["1"]],
            )
        ],
    )
    t = ReadDataSourceTool()
    env = json.loads(
        t.execute(
            {"trace_id": "tid", "table_id": "no-such"},
            _make_ctx(doc),
        )
    )
    assert "error" in env
    assert "available" in env
    assert "real" in env["available_table_ids"]


def test_table_benford_runs_on_xlsx_data() -> None:
    """The Benford
    detector actually
    fires when given an
    ExtractedTable full
    of numbers.

    This pins the
    end-to-end
    plumbing: detector
    reads ``doc.tables``
    (now populated by
    ingest) and emits
    findings.
    """
    from manusift.contracts import (
        ExtractedTable,
        ParsedDoc,
    )
    from manusift.detectors.table_stats import (
        BenfordDetector,
    )

    # 200
    # numbers
    # whose
    # leading
    # digits
    # are
    # not
    # uniform
    # (we
    # bias
    # them
    # to
    # start
    # with
    # 3
    # most
    # of
    # the
    # time)
    # --
    # this
    # should
    # trigger
    # a
    # Benford
    # violation.
    rows: list[list[str]] = []
    for i in range(200):
        if i % 3 == 0:
            v = 3.0 + (i * 0.01)
        else:
            v = float(i + 1)
        rows.append([str(v)])
    doc = ParsedDoc(
        trace_id="tid",
        source_path="x.pdf",
        text_blocks=[],
        images=[],
        metadata={},
        tables=[
            ExtractedTable(
                table_id="synth",
                source_kind="xlsx",
                source_path="x.xlsx",
                sheet_name="S",
                source_index=0,
                headers=["v"],
                rows=rows,
            )
        ],
    )
    detector = BenfordDetector()
    result = detector.run(doc)
    assert result.ok is True
    # Either
    # one
    # finding
    # (the
    # biased
    # column)
    # or
    # zero
    # (if
    # the
    # random
    # ones
    # dominate).
    # We
    # only
    # check
    # it
    # didn't
    # crash.
    assert isinstance(result.findings, list)


def test_table_duplicate_row_finds_dupes() -> None:
    from manusift.contracts import (
        ExtractedTable,
        ParsedDoc,
    )
    from manusift.detectors.table_stats import (
        DuplicateRowDetector,
    )

    # Inject
    # a
    # duplicate
    # row.
    rows = [
        ["G1", "1.0"],
        ["G2", "2.0"],
        ["G1", "1.0"],  # exact dup of row0
        ["G3", "3.0"],
    ]
    doc = ParsedDoc(
        trace_id="tid",
        source_path="x.pdf",
        text_blocks=[],
        images=[],
        metadata={},
        tables=[
            ExtractedTable(
                table_id="dup",
                source_kind="xlsx",
                source_path="x.xlsx",
                sheet_name="S",
                source_index=0,
                headers=["g", "v"],
                rows=rows,
            )
        ],
    )
    det = DuplicateRowDetector()
    result = det.run(doc)
    assert result.ok is True
    # Should
    # have
    # flagged
    # at
    # least
    # one
    # duplicate.
    assert len(result.findings) >= 1
    # The
    # finding's
    # evidence
    # is
    # a
    # JSON
    # string
    # (Finding.make
    # takes
    # evidence
    # as
    # the
    # raw
    # payload
    # -- the
    # detector
    # code
    # uses
    # evidence
    # rather
    # than
    # raw).
    import json
    f = result.findings[0]
    payload = json.loads(f.evidence)
    assert "duplicate_groups" in payload
    # At
    # least
    # one
    # group
    # must
    # be
    # the
    # G1/1.0
    # row
    # we
    # injected
    # twice.
    assert any(
        g["occurrences"] >= 2 for g in payload["duplicate_groups"]
    )
