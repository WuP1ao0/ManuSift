"""Companion-data-file ingest (R-audit, 2026-06).

The PDF is not the whole
story. Many manuscripts
ship their raw numbers in
companion XLSX / CSV / TSV
files: ``Source_Data_Fig1.xlsx``,
``MOESM3_supplementary_data.xlsx``,
``data.csv`` and so on.
The four table-statistics
detectors (``BenfordDetector``,
``DuplicateRowDetector``,
``OutlierDetector``,
``RoundBiasDetector``) all
read ``doc.tables`` -- and
before this module existed
``doc.tables`` was always
empty because nothing ever
populated it.

This module turns a path
into a list of
``ExtractedTable`` records.
It supports three input
formats:

  1. ``.xlsx`` -- via
     ``openpyxl`` (3.1.x).
     Each sheet becomes one
     ``ExtractedTable``. Empty
     sheets and sheets with
     no header row are
     skipped.

  2. ``.csv`` / ``.tsv`` --
     ``csv`` from the standard
     library. Each file
     becomes one
     ``ExtractedTable``.

  3. ``.json`` -- a
     ``{"headers": [...],
     "rows": [[...]]}`` shape.
     Useful when the user has
     already exported data via
     the ``extract_table_from_image``
     tool.

The module is deliberately
*filesystem-only*. Callers
point it at a path and get
back records; it does not
decide *which* files to read
or *when*. The web upload
flow uses
``ingest.discover_companion_files``
to find the right files for a
given job, then calls
``parse_data_file`` on each
one. CLI callers use the
public ``iter_data_files_in``
helper.
"""
from __future__ import annotations

import csv
import hashlib
import io
import logging
from pathlib import Path, PurePosixPath
from typing import Any, Iterable
import zipfile

from ..contracts import ExtractedTable

log = logging.getLogger(__name__)


# File extensions we
# know how to parse.
# Anything else is
# silently skipped --
# the caller's loop
# iterates many
# files and the
# vast majority are
# images / PDFs we
# do not want to
# false-positive on.
SUPPORTED_EXTENSIONS: frozenset[str] = frozenset(
    {".xlsx", ".csv", ".tsv", ".json"}
)
ARCHIVE_EXTENSIONS: frozenset[str] = frozenset({".zip"})
MAX_ARCHIVE_MEMBER_BYTES = 25 * 1024 * 1024


def _cell_text(cell: object) -> str:
    value = getattr(cell, "value", None)
    if value is None:
        return ""
    return str(value)


def _cell_fill(cell: object) -> str | None:
    """Return a compact fill marker for visibly highlighted cells."""
    fill = getattr(cell, "fill", None)
    if fill is None or getattr(fill, "fill_type", None) in (None, "none"):
        return None
    color = getattr(fill, "fgColor", None) or getattr(fill, "start_color", None)
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str):
        normalized = rgb.upper()
        if normalized in {"00000000", "000000"}:
            return None
        return normalized[-6:] if len(normalized) == 8 else normalized
    indexed = getattr(color, "indexed", None)
    if indexed is not None:
        return f"indexed:{indexed}"
    theme = getattr(color, "theme", None)
    if theme is not None:
        return f"theme:{theme}"
    return None


def _highlight_entry(
    cell: object,
    *,
    row: int,
    col: int,
    value: str,
) -> dict[str, object] | None:
    fill = _cell_fill(cell)
    if fill is None:
        return None
    return {
        "row": row,
        "col": col,
        "source_row": getattr(cell, "row", 0),
        "source_col": getattr(cell, "column", 0),
        "value": value,
        "fill": fill,
    }


def _short_id(*parts: str) -> str:
    """Stable, short
    identifier for a
    table -- the
    renderer's
    ``markdown_path``
    uses this so the
    LLM can quote
    a specific table
    in its report
    without
    re-discovering
    it.

    Hash the joined
    parts (file path
    + sheet name +
    index) with the
    first 12 hex
    chars of
    ``sha1``. Avoids
    PII in the
    table id while
    still being
    reproducible
    across runs.
    """
    h = hashlib.sha1("|".join(parts).encode("utf-8")).hexdigest()
    return h[:12]


def parse_xlsx(
    path: str | Path,
    *,
    source_path: str | None = None,
) -> list[ExtractedTable]:
    """Turn an ``.xlsx`` file
    into a list of
    ``ExtractedTable``
    (one per non-empty
    sheet, or
    **one per
    fig** if a
    sheet
    contains
    multiple
    fig
    panels).

    OpenPyXL is imported
    lazily so the rest of
    the package loads even
    if ``openpyxl`` is not
    installed -- the
    upload endpoint still
    works for PDF-only
    jobs.

    **R-2026-06-19 (Phase
    C, per-fig
    xlsx):** a sheet like
    ``Sfig.2`` containing
    6 side-by-side figs
    (``Fig.S1a`` ...
    ``Fig.S1f``) now
    produces **6
    ``ExtractedTable``
    records** instead of
    one big flattened
    table. Each
    ``ExtractedTable``
    carries the matched
    fig name in
    ``fig_name`` and the
    (top, bottom, left, right)
    bbox in ``bbox`` so
    detectors can emit
    fig-aware finding
    titles and the user
    can pass precise
    ranges to a
    detector.

    Sheets with **no
    detected fig
    headers** (the
    common single-table
    case) still produce
    **one
    ``ExtractedTable``**
    with the full sheet
    as before.

    Empty sheets (zero rows
    after the header row)
    are skipped. Sheets
    where the header row is
    entirely ``None`` are
    also skipped (they
    usually indicate an
    image-only sheet).
    """
    import openpyxl  # type: ignore

    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(str(p))
    out: list[ExtractedTable] = []
    sp = source_path or str(p)
    # ``data_only=True``
    # means we read
    # computed values
    # (i.e. what
    # Excel would show
    # the user, not the
    # raw formulas). For
    # Benford-style
    # analysis this is
    # what we want;
    # raw formulas
    # would have to be
    # re-evaluated.
    # Note: we do NOT use
    # ``read_only=True``
    # because the
    # fig-boundary
    # detector needs
    # ``ws.max_row`` /
    # ``ws.max_column`` /
    # ``ws.cell(r, c).value``
    # which are not
    # available on
    # ``ReadOnlyWorksheet``.
    wb = openpyxl.load_workbook(
        str(p), data_only=True
    )
    try:
        for s_index, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            try:
                sheet_tables = _parse_xlsx_sheet(
                    ws, sp, sheet_name, s_index
                )
            except Exception as exc:  # noqa: BLE001
                # Malformed sheet -- skip and
                # continue. Other sheets in
                # the workbook may still be
                # useful. The bare exception
                # is logged by the caller
                # via the trace logger.
                import logging
                logging.getLogger(__name__).info(
                    "xlsx sheet %r failed: %s",
                    sheet_name,
                    exc,
                )
                continue
            out.extend(sheet_tables)
    finally:
        wb.close()
    return out


def _parse_xlsx_sheet(
    ws: Any,
    sp: str,
    sheet_name: str,
    s_index: int,
) -> list[ExtractedTable]:
    """Parse one xlsx worksheet into one or more ``ExtractedTable``.

    R-2026-06-19 (Phase C):
    splits the sheet into
    one ``ExtractedTable``
    per fig (when multiple
    fig headers are
    detected), or one
    ``ExtractedTable`` for
    the whole sheet when
    no fig headers are
    found.

    The bbox stored on each
    ``ExtractedTable`` uses
    0-indexed
    [top..bottom] rows
    (inclusive) and
    [left..right] columns
    (right is exclusive on
    the slice but
    inclusive in the
    stored ``bbox`` dict
    so the renderer can
    show "rows 1-6, cols
    1-3" without +1
    arithmetic).
    """
    out: list[ExtractedTable] = []
    from ..tools.safe_read_b import detect_xlsx_figs

    bboxes = detect_xlsx_figs(ws)
    if not bboxes:
        # No fig headers --
        # emit the sheet as
        # one table (legacy
        # behavior, preserves
        # the single-table
        # case).
        _append_sheet_as_one_table(
            ws, sp, sheet_name, s_index, out
        )
        return out

    if len(bboxes) == 1:
        # One fig header but
        # the sheet also has
        # data below it -- a
        # "single-fig sheet".
        # Emit one table for
        # the whole sheet
        # but tag it with
        # ``fig_name`` so the
        # detector title can
        # still mention the
        # fig.
        bb = bboxes[0]
        _append_sheet_as_one_table(
            ws,
            sp,
            sheet_name,
            s_index,
            out,
            fig_name=bb["name"],
            bbox={
                "top": bb["top"],
                "bottom": bb["bottom"],
                "left": bb["left"],
                "right": bb["right"],
            },
        )
        return out

    # Multi-fig sheet:
    # emit one
    # ``ExtractedTable``
    # per fig, each
    # bounded to its
    # (rows, cols) bbox.
    for bb in bboxes:
        # Extract the header
        # row: the row
        # containing the
        # fig name is the
        # header, but
        # ``bbox.top`` ==
        # ``header_row`` so
        # the first row of
        # the bbox is the
        # header. The data
        # rows are
        # ``top+1..bottom``.
        headers: list[str] = []
        rows: list[list[str]] = []
        highlighted_cells: list[dict[str, object]] = []
        for r in range(bb["top"], bb["bottom"] + 1):
            row_cells: list[str] = []
            row_highlights: list[dict[str, object]] = []
            for c in range(bb["left"], bb["right"] + 1):
                cell = ws.cell(row=r + 1, column=c + 1)
                value = _cell_text(cell)
                row_cells.append(value)
                entry = _highlight_entry(
                    cell,
                    row=len(rows),
                    col=c - bb["left"],
                    value=value,
                )
                if entry is not None:
                    row_highlights.append(entry)
            if r == bb["top"]:
                # If the
                # first
                # row
                # looks
                # like a
                # header
                # (mostly
                # text,
                # not
                # numbers),
                # use it
                # as the
                # table
                # header.
                # Otherwise
                # use
                # generic
                # col_n
                # headers
                # so the
                # detector
                # can
                # still
                # reference
                # columns.
                non_empty = [
                    c for c in row_cells if c.strip()
                ]
                if non_empty and all(
                    self_str_is_label(c)
                    for c in non_empty[:3]
                ):
                    headers = row_cells
                else:
                    headers = [
                        f"col_{i + 1}"
                        for i in range(len(row_cells))
                    ]
                    rows.append(row_cells)
                    highlighted_cells.extend(row_highlights)
            else:
                # Skip
                # fully
                # empty
                # rows.
                if not any(c.strip() for c in row_cells):
                    continue
                rows.append(row_cells)
                highlighted_cells.extend(row_highlights)
        # Skip
        # bboxes
        # that
        # produced
        # zero
        # data
        # rows
        # (e.g.
        # a
        # header-only
        # column
        # that's
        # just
        # a
        # legend).
        if not rows:
            continue
        out.append(
            ExtractedTable(
                table_id=_short_id(
                    sp, sheet_name, str(s_index), bb["name"],
                ),
                source_kind="xlsx",
                source_path=sp,
                sheet_name=sheet_name,
                source_index=s_index,
                headers=headers,
                rows=rows,
                fig_name=bb["name"],
                bbox={
                    "top": bb["top"],
                    "bottom": bb["bottom"],
                    "left": bb["left"],
                    "right": bb["right"],
                },
                highlighted_cells=highlighted_cells,
            )
        )
    return out


def _append_sheet_as_one_table(
    ws: Any,
    sp: str,
    sheet_name: str,
    s_index: int,
    out: list[ExtractedTable],
    *,
    fig_name: str = "",
    bbox: dict[str, int] | None = None,
) -> None:
    """Append the whole sheet as one ``ExtractedTable`` (legacy / single-fig path).

    R-2026-06-19 (Phase C):
    preserves the original
    ``parse_xlsx`` behavior
    for sheets with zero or
    one detected fig --
    one ``ExtractedTable``
    covering the entire
    sheet. When ``fig_name``
    is non-empty (one fig
    header detected) the
    table is tagged with
    the fig name + bbox
    so detector titles can
    still say "Fig.1
    column X" instead of
    "Table 1 column X".
    """
    headers: list[str] = []
    rows: list[list[str]] = []
    highlighted_cells: list[dict[str, object]] = []
    for i, row in enumerate(ws.iter_rows()):
        cells = [_cell_text(cell) for cell in row]
        if i == 0:
            headers = cells
            continue
        if not any(c.strip() for c in cells):
            continue
        table_row = len(rows)
        for j, cell in enumerate(row):
            entry = _highlight_entry(
                cell,
                row=table_row,
                col=j,
                value=cells[j],
            )
            if entry is not None:
                highlighted_cells.append(entry)
        rows.append(cells)
    if not rows:
        return
    if not any(h.strip() for h in headers):
        return
    out.append(
        ExtractedTable(
            table_id=_short_id(
                sp, sheet_name, str(s_index), fig_name,
            ),
            source_kind="xlsx",
            source_path=sp,
            sheet_name=sheet_name,
            source_index=s_index,
            headers=headers,
            rows=rows,
            fig_name=fig_name,
            bbox=bbox,
            highlighted_cells=highlighted_cells,
        )
    )


def self_str_is_label(s: str) -> bool:
    """Return True if ``s`` looks like a column header (text) not a data cell.

    R-2026-06-19 (Phase C):
    heuristic used by
    ``_parse_xlsx_sheet`` to
    decide whether the first
    row of a fig bbox is a
    header (use as
    ``ExtractedTable.headers``)
    or already a data row
    (wrap with generic
    ``col_N`` headers so the
    detector has something
    to reference).

    The heuristic: a string
    "looks like a label" if
    it (a) cannot be parsed
    as a float and (b) is
    short (â‰¤ 40 chars) and
    (c) contains at least one
    non-digit character.
    """
    s = s.strip()
    if not s or len(s) > 40:
        return False
    try:
        float(s)
        return False
    except ValueError:
        pass
    return any(not c.isdigit() for c in s)


def parse_csv(
    path: str | Path,
    *,
    delimiter: str | None = None,
    source_path: str | None = None,
) -> list[ExtractedTable]:
    """Turn a CSV / TSV
    file into one
    ``ExtractedTable``.

    The delimiter is
    auto-detected (``csv.Sniffer``)
    if not provided. If the
    sniffer fails -- or
    returns an ambiguous
    single-column result on
    a file that visually
    contains commas -- we
    fall back to
    ``","`` so a manual
    retry is only needed
    for genuinely exotic
    separators.
    The caller can force
    ``"\\t"`` via
    ``delimiter=...`` for
    TSV files that confuse
    the sniffer.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(str(p))
    sp = source_path or str(p)
    with open(p, "r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(8192)
        f.seek(0)
        resolved_delimiter: str | None = delimiter
        if resolved_delimiter is None:
            try:
                dialect = csv.Sniffer().sniff(
                    sample, delimiters=",;\t|"
                )
                resolved_delimiter = dialect.delimiter
            except csv.Error:
                resolved_delimiter = ","
        # Validate
        # the
        # sniffer's
        # choice:
        # if
        # the
        # first
        # row
        # parses
        # as
        # exactly
        # one
        # cell
        # while
        # the
        # delimiter
        # appears
        # in
        # the
        # raw
        # sample,
        # the
        # sniffer
        # is
        # wrong.
        if delimiter is None:
            first_guess = next(
                csv.reader(
                    io.StringIO(sample),
                    delimiter=resolved_delimiter,
                ),
                None,
            )
            if (
                first_guess is not None
                and len(first_guess) == 1
                and resolved_delimiter in sample
            ):
                # Try
                # the
                # other
                # common
                # delimiters.
                for alt in (",", "\t", ";", "|"):
                    if alt == resolved_delimiter:
                        continue
                    alt_guess = next(
                        csv.reader(
                            io.StringIO(sample),
                            delimiter=alt,
                        ),
                        None,
                    )
                    if (
                        alt_guess is not None
                        and len(alt_guess) > 1
                    ):
                        resolved_delimiter = alt
                        break
        reader = csv.reader(
            f, delimiter=resolved_delimiter
        )
        rows_iter = list(reader)
    if not rows_iter:
        return []
    headers = rows_iter[0]
    rows: list[list[str]] = []
    for r in rows_iter[1:]:
        cells = [(c or "").strip() for c in r]
        if not any(cells):
            continue
        rows.append(cells)
    if not rows:
        return []
    return [
        ExtractedTable(
            table_id=_short_id(sp, "", "0"),
            source_kind="csv",
            source_path=sp,
            sheet_name="",
            source_index=0,
            headers=[h.strip() for h in headers],
            rows=rows,
        )
    ]


def parse_json_table(
    path: str | Path,
    *,
    source_path: str | None = None,
) -> list[ExtractedTable]:
    """Read a JSON file of
    the form
    ``{"headers": [...],
    "rows": [[...]]}`` and
    return one
    ``ExtractedTable``.

    This is the format the
    ``extract_table_from_image``
    tool emits. Reusing the
    same shape means the
    table-statistics
    detectors can analyse
    OCR'd data without
    changing their input
    contract.
    """
    import json

    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(str(p))
    sp = source_path or str(p)
    payload = json.loads(p.read_text(encoding="utf-8"))
    headers = payload.get("headers") or []
    rows_raw = payload.get("rows") or []
    rows: list[list[str]] = []
    for r in rows_raw:
        if not isinstance(r, list):
            continue
        cells = ["" if c is None else str(c) for c in r]
        if not any(c.strip() for c in cells):
            continue
        rows.append(cells)
    if not rows:
        return []
    return [
        ExtractedTable(
            table_id=_short_id(sp, "", "0"),
            source_kind="json",
            source_path=sp,
            sheet_name="",
            source_index=0,
            headers=[str(h) for h in headers],
            rows=rows,
        )
    ]


def parse_data_file(
    path: str | Path,
) -> list[ExtractedTable]:
    """Parse a single file
    by extension.

    Returns an empty list
    if the extension is not
    one of
    ``SUPPORTED_EXTENSIONS``.
    Raises
    ``FileNotFoundError``
    if the path does not
    exist and ``ValueError``
    if openpyxl is not
    installed but the file
    is an .xlsx (so the
    caller can surface a
    helpful error to the
    user rather than a
    bare ``ModuleNotFoundError``).
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(str(p))
    ext = p.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        return []
    if ext == ".xlsx":
        try:
            return parse_xlsx(p)
        except ModuleNotFoundError as exc:
            raise ValueError(
                "XLSX parsing requires the openpyxl package; "
                "run `pip install openpyxl`."
            ) from exc
    if ext == ".csv":
        return parse_csv(p)
    if ext == ".tsv":
        return parse_csv(p, delimiter="\t")
    if ext == ".json":
        return parse_json_table(p)
    # unreachable
    return []


def _safe_archive_base(zip_path: Path) -> str:
    stem = "".join(
        c if c.isalnum() or c in ("_", "-") else "_"
        for c in zip_path.stem
    )
    return f"{stem}_{_short_id(str(zip_path.resolve()))}"


def _iter_zip_data_files(
    zip_path: Path,
    extract_root: Path,
) -> Iterable[Path]:
    dest_root = extract_root / _safe_archive_base(zip_path)
    try:
        archive = zipfile.ZipFile(zip_path)
    except zipfile.BadZipFile:
        log.info(
            "supplementary archive is not a valid zip",
            extra={"file": str(zip_path)},
        )
        return
    with archive:
        for info in archive.infolist():
            if info.is_dir():
                continue
            member = PurePosixPath(info.filename.replace("\\", "/"))
            if (
                member.is_absolute()
                or ".." in member.parts
                or member.suffix.lower() not in SUPPORTED_EXTENSIONS
            ):
                continue
            if info.file_size > MAX_ARCHIVE_MEMBER_BYTES:
                log.info(
                    "supplementary archive member skipped as too large",
                    extra={
                        "file": str(zip_path),
                        "member": info.filename,
                        "size": info.file_size,
                    },
                )
                continue
            target = dest_root.joinpath(*member.parts)
            target.parent.mkdir(parents=True, exist_ok=True)
            with archive.open(info) as src, target.open("wb") as dst:
                while True:
                    chunk = src.read(1024 * 1024)
                    if not chunk:
                        break
                    dst.write(chunk)
            yield target


def iter_data_files_in(
    root: str | Path,
    *,
    max_depth: int = 3,
    extract_archives_to: str | Path | None = None,
) -> Iterable[Path]:
    """Yield every supported
    data file under
    ``root`` up to
    ``max_depth`` levels
    deep.

    The companion-file
    discovery uses this
    helper to enumerate
    every XLSX / CSV /
    TSV / JSON it can find
    without recursing
    indefinitely. Hidden
    directories (``\\.name``)
    are skipped so the
    helper does not pick
    up ``.git`` /
    ``__pycache__`` etc.
    """
    root_p = Path(root)
    if not root_p.exists():
        return
    base_depth = len(root_p.resolve().parts)
    for p in root_p.rglob("*"):
        if p.is_dir():
            continue
        # Skip
        # hidden
        # dirs
        # by
        # looking
        # at
        # any
        # parent
        # part.
        try:
            depth = len(p.resolve().parts) - base_depth
        except OSError:
            continue
        if depth > max_depth:
            continue
        if any(
            part.startswith(".")
            for part in p.relative_to(root_p).parts[:-1]
        ):
            continue
        suffix = p.suffix.lower()
        if suffix in SUPPORTED_EXTENSIONS:
            yield p
            continue
        if suffix in ARCHIVE_EXTENSIONS and extract_archives_to is not None:
            yield from _iter_zip_data_files(
                p, Path(extract_archives_to)
            )


def discover_companion_files(
    upload_dir: str | Path,
    *,
    extract_archives_to: str | Path | None = None,
) -> list[Path]:
    """Find every XLSX /
    CSV / TSV / JSON in the
    user's upload
    directory.

    The web upload endpoint
    stores companion files
    in ``<job_dir>/materials/``
    (see the upload
    handler in
    ``manusift/web/app.py``);
    the CLI expects them in
    the same dir as the
    PDF. This helper
    abstracts over both.
    """
    return list(
        iter_data_files_in(
            upload_dir,
            extract_archives_to=extract_archives_to,
        )
    )
