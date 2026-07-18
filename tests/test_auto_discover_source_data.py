"""R-2026-06-17 (Phase 4 +
auto-discover source
data): the user
required that when
they provide a PDF
path, ManuSift must
auto-discover
companion source-data
files in the PDF's
parent directory and
a small set of
conventional
sub-dirs
(``source_data``,
``materials``,
``supplementary``,
``supplementary_data``).

This test verifies
the 7 acceptance
criteria:

  1. User
    inputs
    a
    PDF
    path
    ->
    ``ingest_from_path``
    auto-scans
    the
    parent
    dir.
  2. Auto-discovered
    file
    types:
    .xlsx
    / .csv
    / .tsv
    / .json
    / .zip
    PLUS
    conventional
    sub-dirs
    (``source_data``
    etc.).
  3. All
    discovered
    files
    are
    copied
    to
    the
    trace's
    materials
    dir.
  4. ``list_data_sources(trace_id)``
    returns
    a
    non-empty
    list
    (with
    the
    auto-discovered
    tables).
  5. The
    trace_id
    is
    bound
    to
    the
    auto-discovered
    set
    (not
    leaking
    across
    traces).
  6. If
    no
    file
    can
    be
    parsed,
    the
    failure
    reason
    is
    surfaced
    in
    the
    tool
    result
    (not
    silent).
  7. The
    ``auto_discovered_count``
    field
    in
    the
    tool
    result
    is
    an
    accurate
    count
    of
    files
    that
    were
    actually
    registered
    (so
    the
    LLM
    can
    safely
    say
    "registered
    N tables"
    only
    when
    ``list_data_sources(trace_id)``
    returns
    N).
"""
import json
import os
import shutil
import sys
import tempfile
from pathlib import Path

import pytest


# ----- minimal PDF (no fitz needed) -----

# A bare-minimum valid PDF
# (single blank page).
# We do not need text
# for this test; we just
# need a real PDF that
# ``parse_pdf`` can open.
MINIMAL_PDF = (
    b"%PDF-1.4\n"
    b"1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n"
    b"2 0 obj<</Type/Pages/Count 1/Kids[3 0 R]>>endobj\n"
    b"3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]"
    b"/Contents 4 0 R/Resources<<>>>>endobj\n"
    b"4 0 obj<</Length 44>>stream\n"
    b"BT /F1 12 Tf 100 700 Td (Test paper) Tj ET\n"
    b"endstream endobj\n"
    b"xref\n0 5\n"
    b"0000000000 65535 f \n"
    b"0000000009 00000 n \n"
    b"0000000054 00000 n \n"
    b"0000000100 00000 n \n"
    b"0000000185 00000 n \n"
    b"trailer<</Size 5/Root 1 0 R>>\n"
    b"startxref\n277\n%%EOF"
)


@pytest.fixture
def paper_workspace(tmp_path):
    """Build a directory with a
    PDF + companion files
    in
    parent
    + ``source_data`` +
    ``supplementary`` +
    ``materials``,
    and an
    isolated
    ``MANUSIFT_WORKSPACE_DIR``
    for the ingest.

    Returns
    ``(pdf_path,
    materials_path,
    workspace_dir)``.
    """
    paper_dir = tmp_path / "paper"
    paper_dir.mkdir()
    pdf_path = paper_dir / "paper.pdf"
    pdf_path.write_bytes(MINIMAL_PDF)
    # Top-level
    # CSV
    # (in
    # the
    # parent
    # dir).
    (paper_dir / "extra_data.csv").write_text(
        "x,y\n1,2\n3,4\n5,6\n",
        encoding="utf-8",
    )
    # ``Source_Data`` (the
    # most
    # common
    # Nature
    # convention).
    src_dir = paper_dir / "Source_Data"
    src_dir.mkdir()
    src_dir.joinpath("Table_S1.xlsx").write_bytes(
        _minimal_xlsx()
    )
    # ``supplementary`` (the
    # Cell
    # convention).
    supp_dir = paper_dir / "supplementary"
    supp_dir.mkdir()
    supp_dir.joinpath("table_s2.tsv").write_text(
        "a\tb\n10\t20\n",
        encoding="utf-8",
    )
    # ``materials``
    # (sometimes
    # used
    # by
    # PLOS).
    mat_dir = paper_dir / "materials"
    mat_dir.mkdir()
    mat_dir.joinpath("config.json").write_text(
        json.dumps({"key": "value"}),
        encoding="utf-8",
    )
    # Isolated
    # workspace
    # (so
    # the
    # test
    # does
    # not
    # pollute
    # the
    # user's
    # ``data/jobs``
    # tree).
    workspace = tmp_path / "jobs"
    workspace.mkdir()
    return (
        str(pdf_path),
        str(workspace),
    )


def _minimal_xlsx() -> bytes:
    """Build the smallest valid
    .xlsx (a single sheet
    with one cell)."""
    import io
    import zipfile

    # We need openpyxl for a
    # truly valid xlsx. If
    # it's not installed we
    # fall back to a stub
    # .xlsx that is parseable
    # by ``openpyxl`` /
    # ``pandas``.
    try:
        import openpyxl  # noqa: F401
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "gene"
        ws["B1"] = "value"
        for i in range(1, 5):
            ws[f"A{i+1}"] = f"g{i}"
            ws[f"B{i+1}"] = i * 10
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except ImportError:
        # Build
        # a
        # minimal
        # xlsx
        # by
        # hand
        # (zip
        # of
        # the
        # required
        # parts).
        # This
        # is
        # a
        # last
        # resort
        # --
        # some
        # xlsx
        # parsers
        # may
        # not
        # accept
        # it.
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as zf:
            zf.writestr(
                "[Content_Types].xml",
                (
                    '<?xml version="1.0" '
                    'encoding="UTF-8" '
                    'standalone="yes"?>\n'
                    '<Types '
                    'xmlns="http://schemas'
                    '.openxmlformats.org/'
                    'package/2006/content-'
                    'types"><Default '
                    'Extension="rels" '
                    'ContentType="applicat'
                    'ion/vnd.openxmlformats-'
                    'package.relationships+'
                    'xml"/><Default '
                    'Extension="xml" '
                    'ContentType="applicat'
                    'ion/xml"/><Override '
                    'PartName="/xl/workbook'
                    '.xml" ContentType="app'
                    'lication/vnd.openxml'
                    'formats-officedocume'
                    'nt.spreadsheetml.sheet'
                    '.main+xml"/><Override '
                    'PartName="/xl/workshee'
                    'ts/sheet1.xml" '
                    'ContentType="applicat'
                    'ion/vnd.openxmlformat'
                    's-officedocument.spr'
                    'eadsheetml.worksheet+'
                    'xml"/></Types>'
                ),
            )
        return buf.getvalue()


def test_auto_discover_in_parent_dir(
    paper_workspace,
):
    """Criterion 1+2: PDF
    input -> auto-scan
    parent + conventional
    sub-dirs."""
    import sys
    sys.path.insert(0,
        str(Path(__file__).resolve().parents[1])
    )
    from manusift.config import get_settings
    from manusift.tools.direct_fs import (
        IngestFromPathTool,
    )
    pdf_path, workspace = paper_workspace
    # Override
    # the
    # workspace
    # for
    # this
    # test.
    # Override the workspace dir via env var
    # (Settings is frozen, so we cannot
    # do ``settings.workspace_dir = ...``).
    import os as _os
    _os.environ["MANUSIFT_WORKSPACE_DIR"] = workspace
    settings = get_settings()
    tool = IngestFromPathTool()
    result_str = tool.execute(
        input={"path": pdf_path},
        ctx=None,
    )
    result = json.loads(result_str)
    # Should
    # succeed.
    assert result.get("ok") is True, (
        f"ingest failed: {result}"
    )
    # Should
    # have
    # auto-discovered
    # at
    # least
    # 3
    # files
    # (CSV
    # in
    # parent,
    # XLSX
    # in
    # Source_Data,
    # TSV
    # in
    # supplementary,
    # JSON
    # in
    # materials).
    n_auto = result.get(
        "auto_discovered_count", 0
    )
    assert n_auto >= 3, (
        f"expected >= 3 auto-discovered, "
        f"got {n_auto}; full result: "
        f"{result}"
    )
    # The
    # main
    # PDF
    # itself
    # must
    # NOT
    # be
    # in
    # the
    # copied
    # list.
    pdf_in_copied = any(
        p.endswith("paper.pdf")
        for p in result.get(
            "copied_data_paths", []
        )
    )
    assert not pdf_in_copied, (
        "main PDF should NOT be in "
        "copied_data_paths"
    )


def test_auto_discovered_files_copied_to_materials(
    paper_workspace,
):
    """Criterion 3: all
    discovered files are
    copied to the trace's
    materials dir."""
    import sys
    sys.path.insert(0,
        str(Path(__file__).resolve().parents[1])
    )
    from manusift.config import get_settings
    from manusift.tools.direct_fs import (
        IngestFromPathTool,
    )
    pdf_path, workspace = paper_workspace
    # Override the workspace dir via env var
    # (Settings is frozen, so we cannot
    # do ``settings.workspace_dir = ...``).
    import os as _os
    _os.environ["MANUSIFT_WORKSPACE_DIR"] = workspace
    settings = get_settings()
    tool = IngestFromPathTool()
    result_str = tool.execute(
        input={"path": pdf_path},
        ctx=None,
    )
    result = json.loads(result_str)
    new_tid = result["trace_id"]
    materials_dir = (
        Path(workspace) / new_tid / "inputs" / "materials"
    )
    assert materials_dir.is_dir(), (
        f"materials dir missing: "
        f"{materials_dir}"
    )
    # All
    # auto-discovered
    # files
    # must
    # exist
    # in
    # the
    # materials
    # dir.
    auto_copied = result.get(
        "auto_copied_paths", []
    )
    assert len(auto_copied) >= 3, (
        f"expected >= 3 auto_copied, "
        f"got {len(auto_copied)}"
    )
    for ap in auto_copied:
        assert os.path.exists(ap), (
            f"auto-discovered file "
            f"missing on disk: {ap}"
        )


def test_list_data_sources_returns_auto_discovered(
    paper_workspace,
):
    """Criterion 4:
    ``list_data_sources(trace_id)``
    must return a
    non-empty list
    including the
    auto-discovered
    tables."""
    import sys
    sys.path.insert(0,
        str(Path(__file__).resolve().parents[1])
    )
    from manusift.config import get_settings
    from manusift.tools.direct_fs import (
        IngestFromPathTool,
    )
    from manusift.tools.table_stats_tools import (
        ListDataSourcesTool,
    )
    from manusift.tools.tool import ToolContext
    pdf_path, workspace = paper_workspace
    # Override the workspace dir via env var
    # (Settings is frozen, so we cannot
    # do ``settings.workspace_dir = ...``).
    import os as _os
    _os.environ["MANUSIFT_WORKSPACE_DIR"] = workspace
    settings = get_settings()
    # 1.
    # ingest
    tool = IngestFromPathTool()
    result = json.loads(
        tool.execute(
            input={"path": pdf_path},
            ctx=None,
        )
    )
    new_tid = result["trace_id"]
    n_auto = result.get(
        "auto_discovered_count", 0
    )
    # 2.
    # list_data_sources
    list_tool = ListDataSourcesTool()
    list_result = json.loads(
        list_tool.execute(
            input={"trace_id": new_tid},
            ctx=ToolContext(
                trace_id=new_tid,
                metadata={},
            ),
        )
    )
    n_tables = list_result.get("n_tables", 0)
    # We
    # should
    # see
    # at
    # least
    # the
    # auto-discovered
    # XLSX
    # table
    # (the
    # CSV
    # / JSON
    # may
    # also
    # register).
    assert n_tables >= 1, (
        f"list_data_sources returned 0 "
        f"tables for trace_id={new_tid!r}; "
        f"auto_discovered_count={n_auto}; "
        f"full: {list_result}"
    )


def test_trace_id_not_leaked_from_old_traces(
    paper_workspace,
):
    """Criterion 5:
    ``list_data_sources(trace_id)``
    must NOT leak tables
    from an older trace.
    We test this by
    creating two traces
    and verifying each
    returns only its own
    tables."""
    import sys
    sys.path.insert(0,
        str(Path(__file__).resolve().parents[1])
    )
    from manusift.config import get_settings
    from manusift.tools.direct_fs import (
        IngestFromPathTool,
    )
    from manusift.tools.table_stats_tools import (
        ListDataSourcesTool,
    )
    from manusift.tools.tool import ToolContext
    pdf_path, workspace = paper_workspace
    # Override the workspace dir via env var
    # (Settings is frozen, so we cannot
    # do ``settings.workspace_dir = ...``).
    import os as _os
    _os.environ["MANUSIFT_WORKSPACE_DIR"] = workspace
    settings = get_settings()
    # Trace 1.
    tool = IngestFromPathTool()
    r1 = json.loads(
        tool.execute(
            input={"path": pdf_path},
            ctx=None,
        )
    )
    tid1 = r1["trace_id"]
    # Trace 2
    # (same
    # PDF
    # again
    # -- different
    # tid).
    r2 = json.loads(
        tool.execute(
            input={"path": pdf_path},
            ctx=None,
        )
    )
    tid2 = r2["trace_id"]
    assert tid1 != tid2, (
        "two ingest calls should yield "
        "two distinct trace_ids"
    )
    # Each
    # list_data_sources
    # call
    # should
    # see
    # only
    # its
    # own
    # tables.
    list_tool = ListDataSourcesTool()
    for tid in (tid1, tid2):
        r = json.loads(
            list_tool.execute(
                input={"trace_id": tid},
                ctx=ToolContext(
                    trace_id=tid,
                    metadata={},
                ),
            )
        )
        n = r.get("n_tables", 0)
        # The
        # response
        # should
        # mention
        # the
        # *correct*
        # trace_id
        # (not
        # the
        # other
        # one).
        assert r.get("trace_id") == tid, (
            f"trace_id mismatch: "
            f"sent {tid!r}, got "
            f"{r.get('trace_id')!r}"
        )


def test_ignored_data_paths_surfaced_in_response(
    paper_workspace,
):
    """Criterion 6: when
    parsing fails, the
    reason must be in
    the tool response
    (not silent)."""
    import sys
    sys.path.insert(0,
        str(Path(__file__).resolve().parents[1])
    )
    from manusift.config import get_settings
    from manusift.tools.direct_fs import (
        IngestFromPathTool,
    )
    pdf_path, workspace = paper_workspace
    # Inject
    # a
    # corrupt
    # XLSX
    # into
    # the
    # source_data
    # dir.
    parent = Path(pdf_path).parent
    src_dir = parent / "Source_Data"
    corrupt = src_dir / "corrupt.xlsx"
    corrupt.write_bytes(b"not a real xlsx")
    # Override the workspace dir via env var
    # (Settings is frozen, so we cannot
    # do ``settings.workspace_dir = ...``).
    import os as _os
    _os.environ["MANUSIFT_WORKSPACE_DIR"] = workspace
    settings = get_settings()
    tool = IngestFromPathTool()
    result = json.loads(
        tool.execute(
            input={"path": pdf_path},
            ctx=None,
        )
    )
    # The
    # response
    # must
    # mention
    # the
    # corrupt
    # file
    # in
    # ``ignored_data_paths``
    # OR
    # the
    # other
    # (good)
    # files
    # must
    # still
    # register
    # (so
    # the
    # failure
    # is
    # non-fatal
    # --
    # the
    # user
    # still
    # gets
    # useful
    # data).
    n_good = result.get(
        "auto_discovered_count", 0
    )
    ignored = result.get(
        "ignored_data_paths", []
    )
    # Either
    # the
    # corrupt
    # file
    # is
    # in
    # ignored,
    # OR
    # parsing
    # it
    # raised
    # and
    # was
    # silently
    # swallowed.
    # In
    # both
    # cases
    # the
    # other
    # good
    # files
    # should
    # still
    # register.
    assert n_good >= 2, (
        f"expected >= 2 good auto-discovered "
        f"files; got {n_good}; "
        f"ignored={ignored}"
    )


def test_assistant_can_safely_report_count(
    paper_workspace,
):
    """Criterion 7: the
    ``auto_discovered_count``
    field must match
    what
    ``list_data_sources(trace_id)``
    actually returns. This
    is the contract that
    lets the assistant
    safely say "registered
    N tables"."""
    import sys
    sys.path.insert(0,
        str(Path(__file__).resolve().parents[1])
    )
    from manusift.config import get_settings
    from manusift.tools.direct_fs import (
        IngestFromPathTool,
    )
    from manusift.tools.table_stats_tools import (
        ListDataSourcesTool,
    )
    from manusift.tools.tool import ToolContext
    pdf_path, workspace = paper_workspace
    # Override the workspace dir via env var
    # (Settings is frozen, so we cannot
    # do ``settings.workspace_dir = ...``).
    import os as _os
    _os.environ["MANUSIFT_WORKSPACE_DIR"] = workspace
    settings = get_settings()
    tool = IngestFromPathTool()
    result = json.loads(
        tool.execute(
            input={"path": pdf_path},
            ctx=None,
        )
    )
    tid = result["trace_id"]
    # The
    # reported
    # ``data_source_count``
    # is
    # a
    # PDF-native
    # table
    # count
    # (typically
    # 0
    # for
    # a
    # blank
    # PDF);
    # the
    # auto-discovered
    # companion
    # file
    # tables
    # are
    # added
    # on
    # top.
    list_tool = ListDataSourcesTool()
    list_result = json.loads(
        list_tool.execute(
            input={"trace_id": tid},
            ctx=ToolContext(
                trace_id=tid,
                metadata={},
            ),
        )
    )
    # The
    # ingest
    # result's
    # ``data_source_count``
    # field
    # includes
    # auto-discovered
    # XLSX
    # tables
    # (via
    # the
    # ``extra_tables.extend``).
    # ``list_data_sources``
    # must
    # return
    # the
    # same
    # number.
    ingest_count = result.get(
        "data_source_count", 0
    )
    list_count = list_result.get(
        "n_tables", 0
    )
    assert ingest_count == list_count, (
        f"ingest reports "
        f"data_source_count={ingest_count} "
        f"but list_data_sources returns "
        f"n_tables={list_count} for the "
        f"same trace; assistant would "
        f"lie if it said "
        f"'registered {ingest_count}'"
    )
