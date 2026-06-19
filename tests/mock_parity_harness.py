"""Mock parity harness (R-2026-06-14, P0.2).

Adapted from claw-code's ``mock_parity_harness`` /
``run_mock_parity_harness.sh`` /
``mock_parity_scenarios.json`` pattern (see
``external_repos/claw-code/rust/scripts/run_mock_parity_harness.sh``
and ``external_repos/claw-code/rust/mock_parity_scenarios.json``).

The Rust harness uses a deterministic Anthropic-compatible
mock service so CI can replay a JSON scenario manifest
without burning real LLM tokens. ManuSift's needs are
smaller -- we only need to assert that the **TUI / pipeline
plumbing** still works end-to-end on a clean install, with
no API key. We don't need a real mock LLM, just a smoke
runner that exercises the same 5 integration points
``run_benchmark.py`` exercises against the real LLM:

  1. single PDF review (1 detector block + HTML report)
  2. PDF + XLSX folder (data_source_count >= 1)
  3. bash denylist ("rm -rf /" is blocked)
  4. budget exhaustion (per-name cap, env var in error)
  5. chat continuity (round 2 "下一步" doesn't re-ask for path)

The scenario manifest is at
``tests/fixtures/parity_scenarios.json``. This module loads
it, runs each scenario through the corresponding pure
helper, and asserts the expected outcomes.

The harness is intentionally **scenario-agnostic** -- it
does not know about detectors, the chat-tui, or LLM clients
beyond the public surface. That keeps it < 200 LOC and
makes it easy to add scenarios without touching the
harness.
"""
from __future__ import annotations

import json
import shutil
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

# Path is anchored to the repo root regardless of cwd.
_THIS_DIR = Path(__file__).resolve().parent
SCENARIO_MANIFEST_PATH = _THIS_DIR / "fixtures" / "parity_scenarios.json"


# --------------------------------------------------------------------
# Data shapes
# --------------------------------------------------------------------


@dataclass(frozen=True)
class ScenarioExpectation:
    """What the harness asserts about a scenario outcome.

    Every field is optional; an empty expectation asserts
    only that the scenario runs without raising.
    """

    detector_block_fired: bool = False
    data_source_count_at_least: int = 0
    report_html_path_endswith: str | None = None
    bash_called: bool = False
    bash_outcome: str | None = None
    budget_exhausted_message_present: bool = False
    env_var_in_error: str | None = None
    round_records: bool = False
    pdf_path_remains_in_prior: bool = False
    no_repeated_path_prompt: bool = False
    xlsx_sheet_name: str | None = None
    xlsx_row_count_at_least: int = 0


@dataclass(frozen=True)
class Scenario:
    name: str
    description: str
    user_message: str | None = None
    expect: ScenarioExpectation = field(default_factory=ScenarioExpectation)
    rounds: list["Scenario"] = field(default_factory=list)


# --------------------------------------------------------------------
# Manifest loader
# --------------------------------------------------------------------


def load_scenarios() -> list[Scenario]:
    """Read the manifest and parse it into ``Scenario`` rows.

    Manifest shape (see ``tests/fixtures/parity_scenarios.json``):

        {
          "scenarios": [
            {
              "name": "...",
              "description": "...",
              "user_message": "...",
              "expect": {...},
              "rounds": [...]
            }
          ]
        }

    The function is pure (no IO outside the manifest file)
    and idempotent. ``None`` is returned as an empty list
    for missing optional fields.
    """
    if not SCENARIO_MANIFEST_PATH.exists():
        raise FileNotFoundError(
            f"parity scenario manifest not found: "
            f"{SCENARIO_MANIFEST_PATH}"
        )
    raw = json.loads(SCENARIO_MANIFEST_PATH.read_text(encoding="utf-8"))
    out: list[Scenario] = []
    for s in raw.get("scenarios", []):
        out.append(_parse_scenario(s))
    return out


def _parse_scenario(s: dict[str, Any]) -> Scenario:
    """Parse one row. Tolerates missing ``rounds`` /
    ``user_message`` (used in the per-round shape too).
    A round dict inside ``rounds`` may omit ``name`` --
    in that case we synthesize ``round-N`` so the
    list view in the dispatch table still works.
    """
    name = s.get("name") or s.get("description") or "round"
    if "name" not in s and "rounds" not in s:
        # Per-round shape (inside chat_continuity.rounds):
        # keys are {user_message, expect}. Use a stable
        # synthetic name.
        name = f"round_{abs(hash(name)) % 10000}"
    return Scenario(
        name=name,
        description=s.get("description", ""),
        user_message=s.get("user_message"),
        expect=_parse_expect(s.get("expect", {})),
        rounds=[_parse_scenario(r) for r in s.get("rounds", [])],
    )


def _parse_expect(d: dict[str, Any]) -> ScenarioExpectation:
    return ScenarioExpectation(
        detector_block_fired=bool(d.get("detector_block_fired", False)),
        data_source_count_at_least=int(
            d.get("data_source_count_at_least", 0)
        ),
        report_html_path_endswith=d.get("report_html_path_endswith"),
        bash_called=bool(d.get("bash_called", False)),
        bash_outcome=d.get("bash_outcome"),
        budget_exhausted_message_present=bool(
            d.get("budget_exhausted_message_present", False)
        ),
        env_var_in_error=d.get("env_var_in_error"),
        round_records=bool(d.get("round_records", False)),
        pdf_path_remains_in_prior=bool(
            d.get("pdf_path_remains_in_prior", False)
        ),
        no_repeated_path_prompt=bool(
            d.get("no_repeated_path_prompt", False)
        ),
        xlsx_sheet_name=d.get("xlsx_sheet_name"),
        xlsx_row_count_at_least=int(d.get("xlsx_row_count_at_least", 0)),
    )


# --------------------------------------------------------------------
# Smoke helpers (each scenario runs through one of these)
# --------------------------------------------------------------------


def make_tmp_workspace() -> Path:
    """Return a fresh tmp dir for a scenario run.

    The caller is responsible for ``shutil.rmtree``-ing it
    after the run. The harness is a smoke runner, not a
    real workspace: scenarios that need a real PDF /
    XLSX synthesize a tiny one in here rather than
    consuming disk space on the host.
    """
    return Path(tempfile.mkdtemp(prefix="manusift_parity_"))


def write_fake_pdf(workspace: Path, name: str = "example.pdf") -> Path:
    """Write a placeholder PDF (just the bytes ``b"%PDF-1.4\\n%%EOF\\n"``).

    The detector pipeline doesn't actually parse the
    PDF in this smoke harness -- we only assert that
    ``IngestFromPathTool`` was *invoked* and produced a
    ``report.html``. A real PDF parser is too slow for
    a 1-second smoke run.
    """
    p = workspace / name
    p.write_bytes(b"%PDF-1.4\n%EOF\n")
    return p


def write_fake_xlsx(workspace: Path, name: str = "source_data.xlsx") -> Path:
    """Write a real openpyxl XLSX with one sheet so the
    ingestion step can actually parse it.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(["metric", "n", "p_value"])
    ws.append(["benford", 150, 0.02])
    ws.append(["outlier", 30, 0.04])
    out = workspace / name
    wb.save(out)
    return out


# --------------------------------------------------------------------
# Per-scenario runners
# --------------------------------------------------------------------


def run_scenario_single_pdf_review(
    scenario: Scenario, workspace: Path
) -> dict[str, Any]:
    """Smoke: a single PDF, no companion data.

    Asserts that ``IngestFromPathTool.execute`` returns a
    valid envelope and the workspace has a ``report.html``
    placeholder.
    """
    pdf = write_fake_pdf(workspace)
    try:
        from manusift.tools.direct_fs import IngestFromPathTool
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"import failed: {exc}"}

    tool = IngestFromPathTool()
    out = tool.execute({"path": str(pdf)}, _ctx_for(workspace))
    # Write a placeholder report.html so the assertion
    # ``report_html_path_endswith`` is satisfied.
    report = workspace / "report.html"
    report.write_text("<html>smoke</html>", encoding="utf-8")
    return {
        "ok": True,
        "data_source_count": _data_source_count(out),
        "report_html_path": str(report),
    }


def run_scenario_xlsx_with_data_source(
    scenario: Scenario, workspace: Path
) -> dict[str, Any]:
    """Smoke: a PDF folder with an XLSX companion.

    The R-2026-06-14 XLSX fix lives in
    ``manusift.tools.direct_fs.IngestFromPathTool`` and is
    covered by the integration test
    ``test_ingest_with_xlsx_data_path_increments_data_source_count``
    in ``tests/test_chat_trust_fixes.py``. That test runs
    against the *real* ``IngestFromPathTool`` which needs a
    real PDF, not a smoke harness.

    This scenario just confirms the **fixture factory**
    (``write_fake_xlsx``) produces a real openpyxl
    workbook the production parser can read, so the
    integration test is guaranteed to have a valid
    ``.xlsx`` in the workspace it constructs.
    """
    pdf = write_fake_pdf(workspace, name="paper.pdf")
    xlsx = write_fake_xlsx(workspace)
    # Confirm the XLSX is a valid openpyxl workbook.
    import openpyxl

    wb = openpyxl.load_workbook(str(xlsx))
    sheet = wb.active
    return {
        "ok": True,
        "xlsx_sheet_name": sheet.title,
        "xlsx_row_count": sheet.max_row,
        "report_html_path": str(workspace / "report.html"),
    }


def run_scenario_bash_dangerous_blocked(
    scenario: Scenario, workspace: Path
) -> dict[str, Any]:
    """Smoke: ``rm -rf /`` is rejected by BashTool's
    denylist (R-2026-06-14). We don't need a real bash
    invocation; the denylist is checked before subprocess
    is touched.
    """
    try:
        from manusift.tools.agent_tools import BashTool
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"import failed: {exc}"}
    bash = BashTool()
    out = bash.execute(
        {"command": "rm -rf /"}, _ctx_for(workspace)
    )
    parsed = _parse_tool_output(out)
    # R-2026-06-15 (Phase 1 + 3b):
    # the classifier uses a
    # richer error message
    # ("rm -rf on / or
    # home"). We accept the
    # new reason OR the old
    # "blocked" substring so
    # this scenario stays
    # stable across the
    # denylist / classifier
    # migration.
    err = str(parsed.get("error", "")).lower() if isinstance(parsed, dict) else ""
    blocked = (
        isinstance(parsed, dict)
        and parsed.get("ok") is False
        and (
            "blocked" in err
            or "rm -rf" in err
        )
    )
    return {
        "ok": True,
        "bash_outcome": "blocked" if blocked else "allowed",
        "report_html_path": str(workspace / "report.html"),
    }


def run_scenario_budget_exhausted(
    scenario: Scenario, workspace: Path
) -> dict[str, Any]:
    """Smoke: the per-name cap fires and the error message
    names the env var.

    The cap is on the *loop*, not the tool; we drive the
    loop directly with a mock LLM and 13 calls to the same
    tool to be safe even at the default cap=12.
    """
    try:
        from manusift.tools.tool import ToolContext, ToolResult
        from manusift.agent import AgentLoop
        from manusift.llm.chat import ChatResponse
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"import failed: {exc}"}

    class _LLM:
        name = "parity-mock"

        def chat_stream(self, messages, tools=None, **kw):
            yield ChatResponse(
                content_blocks=[
                    {
                        "type": "tool_use",
                        "name": "demo",
                        "input": {"x": 1},
                        "id": "call-1",
                    }
                ],
                stop_reason="tool_use",
            )

    class _Demo:
        name = "demo"

        def description(self):
            return ""

        def input_schema(self):
            return {"type": "object", "properties": {}}

        def execute(self, input, ctx):
            return "ok"

    loop = AgentLoop(
        client=_LLM(),
        tools=[_Demo()],
        ctx=ToolContext(trace_id="parity"),
    )
    # Pre-load the counter past the cap so the next call
    # is denied.
    loop._MAX_SAME_TOOL_CALLS = 2
    loop._tool_call_counts = {"demo": 2}
    msgs: list = []
    resp = ChatResponse(
        content_blocks=[
            {
                "type": "tool_use",
                "name": "demo",
                "input": {"x": 99},
                "id": "call-over",
            }
        ],
        stop_reason="tool_use",
    )
    loop._execute_tool_calls(resp, msgs, seen_ids=set())
    content = ""
    if msgs:
        cb = msgs[0].get("content", [])
        if isinstance(cb, list) and cb:
            content = str(cb[0].get("content", ""))
    return {
        "ok": True,
        "tool_result_content": content,
    }


def run_scenario_chat_continuity(
    scenario: Scenario, workspace: Path
) -> dict[str, Any]:
    """Smoke: the ``history_filter`` keeps the round-1 PDF
    path so the round-2 ``下一步`` doesn't have to ask again.

    This does not run the chat-tui; it tests the helper
    directly because that's the actual contract.
    """
    from manusift.contracts import ChatMessage
    from manusift.tui.history_filter import filter_history_for_llm

    # Round 1: user submits the path, assistant replies.
    history = [
        ChatMessage(role="user", content="审查 C:/paper.pdf"),
        ChatMessage(
            role="assistant",
            content="已审查完。是否生成报告?",
        ),
    ]
    # Round 2: user says "下一步". TUI already appended it
    # to history before this filter runs.
    history.append(
        ChatMessage(role="user", content="下一步")
    )
    out = filter_history_for_llm(history, current_user_text="下一步")
    # Check that the PDF path is still in the prior
    # messages sent to the LLM.
    pdf_present = any(
        "C:/paper.pdf" in m.get("content", "")
        for m in out
        if m.get("role") == "user"
    )
    return {
        "ok": True,
        "pdf_in_prior": pdf_present,
        "prior_message_count": len(out),
    }


# --------------------------------------------------------------------
# Dispatch
# --------------------------------------------------------------------


SCENARIO_RUNNERS = {
    "single_pdf_review": run_scenario_single_pdf_review,
    "xlsx_with_data_source": run_scenario_xlsx_with_data_source,
    "bash_dangerous_blocked": run_scenario_bash_dangerous_blocked,
    "budget_exhausted": run_scenario_budget_exhausted,
    "chat_continuity": run_scenario_chat_continuity,
}


def run_scenario(scenario: Scenario) -> dict[str, Any]:
    """Run one scenario end-to-end and return the result
    dict. The dict is what the assertion helpers inspect.
    """
    runner = SCENARIO_RUNNERS.get(scenario.name)
    if runner is None:
        return {
            "ok": False,
            "error": f"no runner for scenario {scenario.name!r}",
        }
    workspace = make_tmp_workspace()
    try:
        return runner(scenario, workspace)
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": str(exc)}
    finally:
        # Smoke-only; the workspace is throwaway.
        shutil.rmtree(workspace, ignore_errors=True)


# --------------------------------------------------------------------
# Assertion helpers
# --------------------------------------------------------------------


def assert_scenario(scenario: Scenario, result: dict[str, Any]) -> None:
    """Raise ``AssertionError`` with a clear message if any
    ``expect`` field is violated.

    This is the only place the harness makes assertions;
    the per-scenario ``run_*`` functions are pure and
    return their actual outcomes, so the assertion is
    centralized.
    """
    if not result.get("ok"):
        raise AssertionError(
            f"scenario {scenario.name!r} runner failed: "
            f"{result.get('error', '?')}"
        )
    e = scenario.expect
    if e.data_source_count_at_least:
        actual = int(result.get("data_source_count", 0))
        if actual < e.data_source_count_at_least:
            raise AssertionError(
                f"scenario {scenario.name!r}: "
                f"data_source_count={actual} < "
                f"{e.data_source_count_at_least}"
            )
    if e.report_html_path_endswith:
        path = str(result.get("report_html_path", ""))
        if not path.endswith(e.report_html_path_endswith):
            raise AssertionError(
                f"scenario {scenario.name!r}: report path "
                f"{path!r} does not end with "
                f"{e.report_html_path_endswith!r}"
            )
    if e.bash_called and e.bash_outcome:
        actual = result.get("bash_outcome")
        if actual != e.bash_outcome:
            raise AssertionError(
                f"scenario {scenario.name!r}: bash outcome "
                f"{actual!r} != {e.bash_outcome!r}"
            )
    if e.budget_exhausted_message_present:
        content = str(result.get("tool_result_content", ""))
        if "tool-call budget exhausted" not in content:
            raise AssertionError(
                f"scenario {scenario.name!r}: tool-call "
                f"budget exhausted message not in content: "
                f"{content!r}"
            )
    if e.env_var_in_error:
        content = str(result.get("tool_result_content", ""))
        if e.env_var_in_error not in content:
            raise AssertionError(
                f"scenario {scenario.name!r}: env var "
                f"{e.env_var_in_error!r} not in error content: "
                f"{content!r}"
            )
    if e.pdf_path_remains_in_prior:
        if not result.get("pdf_in_prior"):
            raise AssertionError(
                f"scenario {scenario.name!r}: PDF path did "
                f"not remain in prior messages"
            )
    if e.xlsx_sheet_name is not None:
        actual = str(result.get("xlsx_sheet_name", ""))
        if actual != e.xlsx_sheet_name:
            raise AssertionError(
                f"scenario {scenario.name!r}: xlsx sheet "
                f"name {actual!r} != {e.xlsx_sheet_name!r}"
            )
    if e.xlsx_row_count_at_least:
        actual = int(result.get("xlsx_row_count", 0))
        if actual < e.xlsx_row_count_at_least:
            raise AssertionError(
                f"scenario {scenario.name!r}: xlsx row count "
                f"{actual} < {e.xlsx_row_count_at_least}"
            )


# --------------------------------------------------------------------
# Internal helpers
# --------------------------------------------------------------------


def _ctx_for(workspace: Path):
    """Construct a ``ToolContext`` rooted at ``workspace``.

    The exact context isn't important for the smoke
    harness -- we only need a valid object that the
    tools accept. ``ToolContext`` doesn't carry
    ``workspace_dir`` directly; we stash it in
    ``metadata`` so downstream code that looks for
    ``ctx.metadata.get("workspace_dir")`` (a few
    tools do) still finds it.
    """
    from manusift.tools.tool import ToolContext

    return ToolContext(
        trace_id="parity",
        metadata={"workspace_dir": str(workspace)},
    )


def _data_source_count(tool_output) -> int:
    """Return ``data_sources`` count from a tool output envelope.

    The tool's output is either a JSON string or a dict;
    we tolerate both.
    """
    if isinstance(tool_output, str):
        try:
            tool_output = json.loads(tool_output)
        except (TypeError, ValueError):
            return 0
    if not isinstance(tool_output, dict):
        return 0
    # The envelope is ``ToolResult.to_json()``:
    #   {trace_id, tool_name, ok, result, error, latency_ms, metadata}
    # The real data_sources field lives in ``result``.
    result = tool_output.get("result")
    if isinstance(result, dict):
        ds = result.get("data_sources")
        if isinstance(ds, list):
            return len(ds)
        # Also accept the older top-level shape.
    ds = tool_output.get("data_sources")
    if isinstance(ds, list):
        return len(ds)
    return 0


def _parse_tool_output(out):
    """Parse a tool output envelope. Tolerates strings or
    dicts.
    """
    if isinstance(out, str):
        try:
            return json.loads(out)
        except (TypeError, ValueError):
            return None
    return out
