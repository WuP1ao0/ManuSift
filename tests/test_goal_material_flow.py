"""Goal-level material flow regression tests."""
from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import pytest
from PIL import Image

from manusift.contracts import ExtractedImage, ParsedDoc
from manusift.detectors.image_forensics import ImageForensicsDetector
from manusift.detectors.table_relationships import TableRelationshipDetector
from manusift.ingest.xlsx import parse_xlsx
from manusift.report import evidence_builder


def _save_image(img: Image.Image, path: Path) -> ExtractedImage:
    path.parent.mkdir(parents=True, exist_ok=True)
    img.save(path, format="PNG")
    with Image.open(path) as im:
        width, height = im.size
    return ExtractedImage(
        page=0,
        index=0,
        xref=0,
        phash="0" * 16,
        width=width,
        height=height,
        bytes_size=path.stat().st_size,
        exif={},
        image_path=str(path),
    )


def test_synthetic_review_material_flows_to_evidence_report(tmp_path: Path) -> None:
    """One synthetic material set exercises tables, images, and report cards."""
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl.styles import PatternFill

    workbook = tmp_path / "source_data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mixed"
    ws.cell(1, 1, "Fig.3b")
    ws.cell(1, 5, "Fig.3c")
    for offset in (0, 4):
        ws.cell(2, 1 + offset, "Group")
        ws.cell(2, 2 + offset, "A")
        ws.cell(2, 3 + offset, "B")
        ws.cell(2, 4 + offset, "SD")
    rows = [
        ("r1", 10.1, 10.4, 0.1),
        ("r2", 11.2, 11.5, 0.1),
        ("r3", 12.3, 12.6, 0.1),
        ("r4", 13.4, 13.7, 0.1),
        ("r5", 14.5, 14.8, 0.1),
        ("r6", 15.6, 15.9, 0.1),
    ]
    for row_idx, row in enumerate(rows, start=3):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx, value)
        for col_idx, value in enumerate(row, start=5):
            ws.cell(row_idx, col_idx, value)
    ws.cell(4, 6).fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    wb.save(workbook)

    tables = parse_xlsx(workbook)
    by_fig = {table.fig_name: table for table in tables}
    assert {"Fig.3b", "Fig.3c"} <= set(by_fig)
    assert by_fig["Fig.3b"].highlighted_cells == []
    assert by_fig["Fig.3c"].highlighted_cells[0]["fill"] == "FFFF00"

    rng = np.random.default_rng(2028)
    patch_arr = rng.integers(20, 235, (64, 64, 3), dtype=np.uint8)
    patch_arr[16:24, :, :] //= 4
    patch_arr[:, 42:50, :] //= 5
    patch = Image.fromarray(patch_arr)
    img_a = Image.fromarray(
        rng.integers(150, 255, (256, 256, 3), dtype=np.uint8)
    )
    img_b = Image.fromarray(
        rng.integers(0, 120, (256, 256, 3), dtype=np.uint8)
    )
    img_a.paste(patch, (64, 64))
    img_b.paste(patch.rotate(90), (64, 64))
    extracted_a = _save_image(img_a, tmp_path / "fig_a.png")
    extracted_b = _save_image(img_b, tmp_path / "fig_b.png")
    extracted_b = ExtractedImage(
        page=1,
        index=1,
        xref=0,
        phash=extracted_b.phash,
        width=extracted_b.width,
        height=extracted_b.height,
        bytes_size=extracted_b.bytes_size,
        exif={},
        image_path=extracted_b.image_path,
    )

    doc = ParsedDoc(
        trace_id="goal-flow",
        source_path=str(tmp_path / "paper.pdf"),
        text_blocks=[],
        images=[extracted_a, extracted_b],
        metadata={},
        tables=tables,
    )

    table_result = TableRelationshipDetector().run(doc)
    image_result = ImageForensicsDetector().run(doc)
    table_kinds = {finding.raw.get("check") for finding in table_result.findings}
    image_kinds = {finding.raw.get("kind") for finding in image_result.findings}
    assert "fixed_offset" in table_kinds
    assert "zero_variance" in table_kinds
    assert "rotated_texture_overlap" in image_kinds

    findings = [*table_result.findings, *image_result.findings]
    findings_path = tmp_path / "findings.json"
    findings_path.write_text(
        json.dumps(
            {
                "trace_id": "goal-flow",
                "detectors_run": ["table_relationships", "image_forensics"],
                "llm_calls": 0,
                "duration_ms": 0,
                "findings": [
                    {
                        "finding_id": finding.finding_id,
                        "trace_id": finding.trace_id,
                        "detector": finding.detector,
                        "severity": finding.severity,
                        "title": finding.title,
                        "evidence": finding.evidence,
                        "location": finding.location,
                        "raw": finding.raw,
                    }
                    for finding in findings
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    index = evidence_builder.build_evidence_index(
        findings_path=findings_path,
        out_dir=tmp_path / "report",
        paper_id="goal-flow",
    )

    assert any(
        finding.detector == "table_relationships"
        for finding in index.numerical_findings
    )
    assert any(
        finding.detector == "image_forensics"
        for finding in index.visual_findings
    )
    assert (tmp_path / "report" / "data").exists()
