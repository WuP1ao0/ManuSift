"""R-2026-06-19 (Phase D:
per-fig xlsx
+ per-fig
detector run):

End-to-end
tests for the
D-2..D-6 changes
that let the
LLM run a
detector on a
single fig panel.

Tests:

  * D-7a: ``_format_table_label``
    handles all 8
    edge cases (Fig.S1a,
    Fig. S4b, Figure 2,
    Tab.1, no fig,
    no sheet,
    no keyword).
  * D-7b: ``parse_xlsx`` on
    a synthetic 3-fig
    horizontal + 2-fig
    vertical sheet
    produces the
    expected per-fig
    ``ExtractedTable``
    records with
    ``fig_name`` +
    ``bbox``.
  * D-7c: ``BenfordDetector``
    running on a
    per-fig filtered
    doc produces a
    title that
    mentions the
    fig name.
  * D-7d:
    ``DetectorToolAdapter.execute``
    with ``table_ids``
    filter: passes
    through to the
    detector, errors
    on empty / wrong
    type, errors on
    no matches, errors
    on missing tables
    list.
  * D-7e:
    ``ListDataSourcesTool``
    output includes
    ``fig_name`` +
    ``bbox`` for
    per-fig tables.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

import pytest

sys.path.insert(0, r"C:\Users\22509\Desktop\ManuSift1")

from openpyxl import Workbook

from manusift.detectors.table_stats import (
    BenfordDetector,
    _format_table_label,
    _safe_tables,
)
from manusift.ingest.xlsx import parse_xlsx
from manusift.tools.detector_adapter import DetectorToolAdapter
from manusift.tools.tool import ToolContext


# ============================================================================
# D-7a: _format_table_label
# ============================================================================


class TestFormatTableLabel:
    def _label(self, fig_name: str, sheet_name: str = "Sfig.2") -> str:
        class T:
            pass
        t = T()
        t.fig_name = fig_name
        t.sheet_name = sheet_name
        return _format_table_label(t, 0)

    def test_fig_with_dot_prefix(self):
        # "Fig.S1a" should NOT get a redundant "Fig " prefix.
        assert self._label("Fig.S1a") == "Fig.S1a in Sfig.2"

    def test_fig_with_space_prefix(self):
        # "Fig. S4b" (with space) -- keep the space.
        assert self._label("Fig. S4b") == "Fig. S4b in Sfig.2"

    def test_table_prefix(self):
        # "Table S1" should NOT get a redundant "Fig " prefix.
        assert self._label("Table S1") == "Table S1 in Sfig.2"

    def test_figure_prefix(self):
        # "Figure 2" should NOT get a redundant "Fig " prefix.
        assert self._label("Figure 2") == "Figure 2 in Sfig.2"

    def test_tab_prefix(self):
        # "Tab.1" should NOT get a redundant "Fig " prefix.
        assert self._label("Tab.1") == "Tab.1 in Sfig.2"

    def test_no_keyword_gets_fig_prefix(self):
        # "S1a" (no Fig/Table/Tab/Figure prefix) -- add "Fig ".
        assert self._label("S1a") == "Fig S1a in Sfig.2"

    def test_no_fig_falls_back_to_table_label(self):
        # Empty fig_name -- fall back to "Table {sheet} #{n}".
        assert self._label("", sheet_name="S1") == "Table S1 #1"

    def test_no_fig_no_sheet(self):
        # Both empty -- "Table #{n}".
        assert self._label("", sheet_name="") == "Table #1"

    def test_with_suffix(self):
        # The suffix is appended with a space.
        class T:
            pass
        t = T()
        t.fig_name = "Fig.S1a"
        t.sheet_name = "Sfig.2"
        out = _format_table_label(t, 0, suffix="column 'X'")
        assert out == "Fig.S1a in Sfig.2 column 'X'"


# ============================================================================
# D-7b: parse_xlsx per-fig
# ============================================================================


def _build_3h_2v_xlsx(path: Path) -> None:
    """Build a sheet with 3 figs horizontal at top + 2 figs vertical below.

    Layout::

        | Fig.A |       | Fig.B |       | Fig.C |
        |  x  y |       |  x  y |       |  x  y |
        |  1  2 |       |  3  4 |       |  5  6 |
        |  7  8 |       |  9  10|       | 11  12|
        |        <- blank
        | Fig.D |
        |  x  y |
        |  13 14|
        |        <- blank
        | Fig.E |
        |  x  y |
        |  15 16|
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Mixed"
    # Top 3 figs
    ws.cell(1, 1, "Fig.A")
    ws.cell(1, 4, "Fig.B")
    ws.cell(1, 7, "Fig.C")
    for c in (1, 2, 4, 5, 7, 8):
        ws.cell(2, c, "x" if c in (1, 4, 7) else "y")
    for r, vals in enumerate(
        [(1, 2, 3, 4, 5, 6), (7, 8, 9, 10, 11, 12)], start=3
    ):
        for i, v in enumerate(vals, start=1):
            ws.cell(r, i, v)
    # R5 blank
    # Fig.D
    ws.cell(6, 1, "Fig.D")
    ws.cell(7, 1, "x")
    ws.cell(7, 2, "y")
    ws.cell(8, 1, 13)
    ws.cell(8, 2, 14)
    # R9 blank
    # Fig.E
    ws.cell(10, 1, "Fig.E")
    ws.cell(11, 1, "x")
    ws.cell(11, 2, "y")
    ws.cell(12, 1, 15)
    ws.cell(12, 2, 16)
    wb.save(str(path))


def _build_single_table_xlsx(path: Path) -> None:
    """A sheet with NO fig headers -- should emit one ExtractedTable
    with fig_name=""."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Plain"
    ws.cell(1, 1, "Name")
    ws.cell(1, 2, "Age")
    ws.cell(2, 1, "Alice")
    ws.cell(2, 2, 30)
    ws.cell(3, 1, "Bob")
    ws.cell(3, 2, 25)
    wb.save(str(path))


def _build_one_fig_xlsx(path: Path) -> None:
    """A sheet with ONE fig header + lots of data -- emit one table
    with fig_name set."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Single"
    ws.cell(1, 1, "Fig.1")
    ws.cell(2, 1, "x")
    ws.cell(2, 2, "y")
    for r in range(3, 30):
        ws.cell(r, 1, r)
        ws.cell(r, 2, r * 2)
    wb.save(str(path))


class TestParseXlsxPerFig:
    def test_3h_2v_emits_5_tables(self, tmp_path):
        p = tmp_path / "mixed.xlsx"
        _build_3h_2v_xlsx(p)
        tables = parse_xlsx(str(p))
        assert len(tables) == 5
        names = [t.fig_name for t in tables]
        assert names == ["Fig.A", "Fig.B", "Fig.C", "Fig.D", "Fig.E"]

    def test_3h_2v_each_has_bbox(self, tmp_path):
        p = tmp_path / "mixed.xlsx"
        _build_3h_2v_xlsx(p)
        tables = parse_xlsx(str(p))
        for t in tables:
            assert t.bbox is not None
            assert "top" in t.bbox
            assert "bottom" in t.bbox
            assert "left" in t.bbox
            assert "right" in t.bbox

    def test_3h_2v_horizontal_figs_have_short_width(self, tmp_path):
        p = tmp_path / "mixed.xlsx"
        _build_3h_2v_xlsx(p)
        tables = parse_xlsx(str(p))
        # Fig.A, B, C are 3 cols wide (0-2, 4-6, 8-9 inclusive).
        # The right boundary is the last col that belongs to THIS
        # fig (inclusive). Fig.A: left=0, right=2 → 3 cols
        # (0, 1, 2). Fig.B: left=4, right=6 → 3 cols. Fig.C is
        # the rightmost fig and may extend to the last used col.
        a, b, c, d, e = tables
        # R-2026-06-19 (Phase D):
        # bbox now uses inclusive
        # right (``right - 1``
        # of the detector's
        # exclusive right),
        # so width = right - left + 1.
        assert a.bbox["right"] - a.bbox["left"] + 1 == 3
        assert b.bbox["right"] - b.bbox["left"] + 1 == 3
        # 2 vertical figs are 2 cols wide (left=0, right=1).
        assert d.bbox["right"] - d.bbox["left"] + 1 == 2
        assert e.bbox["right"] - e.bbox["left"] + 1 == 2

    def test_3h_2v_vertical_figs_separate_rows(self, tmp_path):
        p = tmp_path / "mixed.xlsx"
        _build_3h_2v_xlsx(p)
        tables = parse_xlsx(str(p))
        d, e = tables[3], tables[4]
        # Fig.D and Fig.E must not overlap in row space.
        assert d.bbox["bottom"] <= e.bbox["top"]

    def test_single_table_no_fig(self, tmp_path):
        p = tmp_path / "plain.xlsx"
        _build_single_table_xlsx(p)
        tables = parse_xlsx(str(p))
        assert len(tables) == 1
        assert tables[0].fig_name == ""
        assert tables[0].bbox is None

    def test_one_fig_with_lots_of_data(self, tmp_path):
        p = tmp_path / "single.xlsx"
        _build_one_fig_xlsx(p)
        tables = parse_xlsx(str(p))
        assert len(tables) == 1
        assert tables[0].fig_name == "Fig.1"
        assert tables[0].bbox is not None

    def test_table_id_includes_fig_name(self, tmp_path):
        p = tmp_path / "mixed.xlsx"
        _build_3h_2v_xlsx(p)
        tables = parse_xlsx(str(p))
        # All 5 table_ids should be unique (different fig names
        # produce different sha1 hashes).
        ids = [t.table_id for t in tables]
        assert len(set(ids)) == 5

    def test_legacy_table_construction_still_works(self):
        """Constructing ExtractedTable without fig_name / bbox should
        still work (backward compat for old callers)."""
        from manusift.contracts import ExtractedTable
        t = ExtractedTable(
            table_id="t1",
            source_kind="csv",
            source_path="/x.csv",
            sheet_name="",
            source_index=0,
            headers=["a", "b"],
            rows=[["1", "2"]],
        )
        assert t.fig_name == ""
        assert t.bbox is None


# ============================================================================
# D-7c: BenfordDetector title with fig_name
# ============================================================================


class FakeFinding:
    def __init__(self, **kwargs):
        self.__dict__.update(kwargs)


def _make_benford_trigger_table(
    fig_name: str = "Fig.S1a",
    sheet_name: str = "Sfig.2",
):
    """Build an ExtractedTable whose Benford chi2 will be < 0.001
    (highly skewed leading digit distribution) so the detector
    fires a finding. Returns the table + the expected title prefix."""
    from manusift.contracts import ExtractedTable
    # 200 values all starting with "1" (leading digit = 1).
    # Benford expects ~30% leading-1, so this is wildly off.
    rows = [[str(1000 + i)] for i in range(200)]
    return ExtractedTable(
        table_id=f"t:{fig_name}",
        source_kind="xlsx",
        source_path="/x.xlsx",
        sheet_name=sheet_name,
        source_index=0,
        headers=["x"],
        rows=rows,
        fig_name=fig_name,
        bbox={"top": 0, "bottom": 200, "left": 0, "right": 1},
    )


class TestBenfordDetectorWithFigName:
    def test_title_mentions_fig_name(self):
        from manusift.contracts import ParsedDoc
        from dataclasses import replace

        t = _make_benford_trigger_table("Fig.S1a", "Sfig.2")
        # Build a ParsedDoc with this table.
        # We need a real ParsedDoc; use the one from contracts.
        from manusift.contracts import TextBlock
        base = ParsedDoc(
            trace_id="trace_d7c",
            source_path="/x.xlsx",
            text_blocks=[],
            images=[],
            metadata={},
            tables=[t],
        )
        d = BenfordDetector()
        result = d.run(base)
        assert result.ok
        # At least one finding with the fig name in the title.
        assert any("Fig.S1a in Sfig.2" in f.title for f in result.findings), (
            f"expected 'Fig.S1a in Sfig.2' in finding titles, got: "
            f"{[f.title for f in result.findings]}"
        )

    def test_legacy_table_uses_table_n_format(self):
        """A table with fig_name="" should produce
        'Table {sheet} #{n} ...' titles."""
        from manusift.contracts import ParsedDoc
        from dataclasses import replace

        t = _make_benford_trigger_table("", "S1")
        base = ParsedDoc(
            trace_id="trace_d7c",
            source_path="/x.xlsx",
            text_blocks=[],
            images=[],
            metadata={},
            tables=[t],
        )
        d = BenfordDetector()
        result = d.run(base)
        assert any("Table S1 #1" in f.title for f in result.findings), (
            f"expected 'Table S1 #1' in titles, got: "
            f"{[f.title for f in result.findings]}"
        )


# ============================================================================
# D-7d: DetectorToolAdapter with table_ids filter
# ============================================================================


def _build_doc_with_3_tables() -> Any:
    """Build a ParsedDoc with 3 distinct tables so we can test
    per-table_id filtering."""
    from manusift.contracts import ParsedDoc, ExtractedTable
    t1 = ExtractedTable(
        table_id="t:1",
        source_kind="xlsx",
        source_path="/x.xlsx",
        sheet_name="S1",
        source_index=0,
        headers=["a"],
        rows=[[str(1000 + i)] for i in range(200)],
        fig_name="Fig.A",
        bbox={"top": 0, "bottom": 1, "left": 0, "right": 1},
    )
    t2 = ExtractedTable(
        table_id="t:2",
        source_kind="xlsx",
        source_path="/x.xlsx",
        sheet_name="S1",
        source_index=0,
        headers=["a"],
        rows=[["2"], ["2"], ["2"]],
        fig_name="Fig.B",
        bbox={"top": 0, "bottom": 1, "left": 0, "right": 1},
    )
    t3 = ExtractedTable(
        table_id="t:3",
        source_kind="xlsx",
        source_path="/x.xlsx",
        sheet_name="S1",
        source_index=0,
        headers=["a"],
        rows=[["3"], ["3"], ["3"]],
        fig_name="Fig.C",
        bbox={"top": 0, "bottom": 1, "left": 0, "right": 1},
    )
    return ParsedDoc(
        trace_id="trace_d7d",
        source_path="/x.xlsx",
        text_blocks=[],
        images=[],
        metadata={},
        tables=[t1, t2, t3],
    )


class TestDetectorToolAdapterTableIdsFilter:
    def _make_ctx(self, doc):
        return ToolContext(
            trace_id=doc.trace_id,
            current_pdf="",
            metadata={"parsed_doc": doc},
        )

    def test_no_table_ids_runs_all(self):
        doc = _build_doc_with_3_tables()
        ctx = self._make_ctx(doc)
        adapter = DetectorToolAdapter(BenfordDetector())
        out = json.loads(adapter.execute({}, ctx))
        assert "findings" in out
        # Without a filter, the detector sees all 3 tables; t:1
        # triggers a Benford violation, t:2/t:3 don't.
        assert any("Fig.A" in f["title"] for f in out["findings"])

    def test_table_ids_filter_scopes_detector(self):
        doc = _build_doc_with_3_tables()
        ctx = self._make_ctx(doc)
        adapter = DetectorToolAdapter(BenfordDetector())
        # Only t:1 should run.
        out = json.loads(
            adapter.execute({"table_ids": ["t:1"]}, ctx)
        )
        # Findings should ONLY mention Fig.A.
        for f in out["findings"]:
            assert "Fig.A" in f["title"] or "Fig.A" in f.get("location", ""), (
                f"filter leaked: {f}"
            )
        # No Fig.B or Fig.C findings.
        assert not any(
            "Fig.B" in f["title"] or "Fig.C" in f["title"]
            for f in out["findings"]
        )

    def test_table_ids_empty_match_returns_error(self):
        doc = _build_doc_with_3_tables()
        ctx = self._make_ctx(doc)
        adapter = DetectorToolAdapter(BenfordDetector())
        out = json.loads(
            adapter.execute({"table_ids": ["nonexistent"]}, ctx)
        )
        assert "error" in out
        assert "no tables matched" in out["error"]
        assert "available" in out
        # The available list should show the 3 real table_ids.
        assert set(out["available"]) == {"t:1", "t:2", "t:3"}

    def test_table_ids_wrong_type_returns_error(self):
        doc = _build_doc_with_3_tables()
        ctx = self._make_ctx(doc)
        adapter = DetectorToolAdapter(BenfordDetector())
        out = json.loads(
            adapter.execute({"table_ids": "t:1"}, ctx)  # string, not list
        )
        assert "error" in out
        assert "must be a list" in out["error"]

    def test_table_ids_does_not_mutate_cached_doc(self):
        doc = _build_doc_with_3_tables()
        ctx = self._make_ctx(doc)
        # Snapshot the doc's table count.
        n_before = len(doc.tables)
        adapter = DetectorToolAdapter(BenfordDetector())
        adapter.execute({"table_ids": ["t:1"]}, ctx)
        # Cached doc must still have all 3 tables.
        assert len(doc.tables) == n_before == 3


# ============================================================================
# D-7e: ListDataSourcesTool with fig_name + bbox
# ============================================================================


class TestListDataSourcesToolFigFields:
    def test_per_fig_table_includes_fig_name_and_bbox(self, tmp_path, monkeypatch):
        """ListDataSourcesTool on a doc with per-fig tables should
        include fig_name + bbox in the output."""
        from manusift.tools.table_stats_tools import ListDataSourcesTool
        from manusift.contracts import ParsedDoc
        from manusift.ingest.xlsx import parse_xlsx

        # Build a synthetic per-fig xlsx.
        p = tmp_path / "mixed.xlsx"
        _build_3h_2v_xlsx(p)
        monkeypatch.chdir(tmp_path)
        monkeypatch.setenv("MANUSIFT_WORKSPACE_DIR", str(tmp_path))

        # Get a real ParsedDoc with these tables.
        tables = parse_xlsx(str(p))
        from dataclasses import replace
        from manusift.contracts import TextBlock
        doc = ParsedDoc(
            trace_id="trace_d7e",
            source_path=str(p),
            text_blocks=[],
            images=[],
            metadata={},
            tables=tables,
        )
        # Patch the tool's doc lookup by using a fake
        # ``_resolve_doc`` that returns our doc.
        from manusift.tools.table_stats_tools import ListDataSourcesTool as _L
        # The tool normally parses the PDF; here we need a
        # doc that has tables. Easiest: directly check the
        # output formatting by running the tool with a custom
        # parse path. But the tool's ``execute`` looks up
        # ``ctx.metadata['parsed_doc']`` first.
        ctx = ToolContext(
            trace_id=doc.trace_id,
            current_pdf="",
            metadata={"parsed_doc": doc},
        )
        out = json.loads(
            _L().execute({"trace_id": doc.trace_id}, ctx)
        )
        assert "tables" in out
        # All 5 tables should have fig_name + bbox.
        for t in out["tables"]:
            assert "fig_name" in t, f"missing fig_name: {t}"
            assert "bbox" in t, f"missing bbox: {t}"
            # bbox is 1-indexed in the output.
            assert t["bbox"]["top"] >= 1
            assert t["bbox"]["left"] >= 1

    def test_legacy_csv_table_has_no_fig_or_bbox(self, tmp_path, monkeypatch):
        """Legacy table (CSV) without fig_name / bbox should NOT have
        those fields in the output (we omit empty fields)."""
        from manusift.tools.table_stats_tools import ListDataSourcesTool as _L
        from manusift.contracts import ParsedDoc, ExtractedTable

        doc = ParsedDoc(
            trace_id="trace_d7e_legacy",
            source_path="/x.csv",
            text_blocks=[],
            images=[],
            metadata={},
            tables=[
                ExtractedTable(
                    table_id="t:csv",
                    source_kind="csv",
                    source_path="/x.csv",
                    sheet_name="",
                    source_index=0,
                    headers=["a", "b"],
                    rows=[["1", "2"]],
                )
            ],
        )
        ctx = ToolContext(
            trace_id=doc.trace_id,
            current_pdf="",
            metadata={"parsed_doc": doc},
        )
        out = json.loads(
            _L().execute({"trace_id": doc.trace_id}, ctx)
        )
        t = out["tables"][0]
        # Legacy table has no fig_name / bbox fields.
        assert "fig_name" not in t
        assert "bbox" not in t


# ============================================================================
# D-7f: Real Nature file end-to-end
# ============================================================================


REAL_FILE = (
    r"C:\Users\22509\ZCodeProject\s41565-025-02082-0"
    r"\Source_Data_MOESM3.xlsx"
)


@pytest.mark.skipif(
    not Path(REAL_FILE).exists(),
    reason="Real Nature SI file not available on this machine",
)
class TestRealNatureFilePerFig:
    def test_parse_emits_per_fig_tables(self):
        tables = parse_xlsx(REAL_FILE)
        # Sfig.2 has 6 figs, Sfig.3 has 2, Sfig.4 has 5.
        by_sheet = {}
        for t in tables:
            by_sheet.setdefault(t.sheet_name, []).append(t)
        if "Sfig.2" in by_sheet:
            assert len(by_sheet["Sfig.2"]) == 6
        if "Sfig.3" in by_sheet:
            assert len(by_sheet["Sfig.3"]) == 2
        if "Sfig.4" in by_sheet:
            assert len(by_sheet["Sfig.4"]) == 5

    def test_benford_on_single_fig_only_runs_on_that_fig(self):
        """End-to-end: build a ParsedDoc from the real file,
        run Benford with table_ids filter scoped to Fig.S1a
        only, and assert no Fig.S1b/c/d/e/f findings appear."""
        from manusift.contracts import ParsedDoc
        from manusift.tools.detector_adapter import DetectorToolAdapter

        tables = parse_xlsx(REAL_FILE)
        sfig2 = [t for t in tables if t.sheet_name == "Sfig.2"]
        if not sfig2:
            pytest.skip("Sfig.2 not in real file")
        fig_s1a = next(
            (t for t in sfig2 if t.fig_name == "Fig.S1a"), None
        )
        if fig_s1a is None:
            pytest.skip("Fig.S1a not in Sfig.2")

        doc = ParsedDoc(
            trace_id="trace_real_filter",
            source_path=REAL_FILE,
            text_blocks=[],
            images=[],
            metadata={},
            tables=sfig2,
        )
        ctx = ToolContext(
            trace_id=doc.trace_id,
            current_pdf="",
            metadata={"parsed_doc": doc},
        )
        adapter = DetectorToolAdapter(BenfordDetector())
        out = json.loads(
            adapter.execute(
                {"table_ids": [fig_s1a.table_id]}, ctx
            )
        )
        # Any findings from this run should mention Fig.S1a,
        # never Fig.S1b/c/d/e/f (the other 5 figs in the same sheet).
        for f in out.get("findings", []):
            title = f.get("title", "")
            for sibling in ("S1b", "S1c", "S1d", "S1e", "S1f"):
                assert sibling not in title, (
                    f"filter leaked: {title} contains {sibling}"
                )
